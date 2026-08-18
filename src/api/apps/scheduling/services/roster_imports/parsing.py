"""Parse uploaded or pasted roster tables into raw preview rows."""

import csv
import io
import zipfile
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

from apps.scheduling.models import RosterImportBatch

from .errors import RosterImportError
from .limits import MAX_COLUMNS, MAX_PREVIEW_ROWS, MAX_UNCOMPRESSED_BYTES, MAX_UPLOAD_BYTES


def _trim_row(values: list) -> list:
    values = list(values)
    while values and (values[-1] is None or values[-1] == ""):
        values.pop()
    return values


def _serialized_cell(cell):
    if getattr(cell, "data_type", "") == "f":
        return {"formula": str(cell.value or "")}
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, int | float | bool | str):
        return value
    return str(value)


def _parse_delimited(payload: bytes, *, worksheet: str) -> list[dict]:
    if len(payload) > MAX_UNCOMPRESSED_BYTES:
        raise RosterImportError("The table exceeds the 25 MiB uncompressed limit.")
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise RosterImportError("CSV and pasted data must be UTF-8 encoded.") from exc
    if "\x00" in text:
        raise RosterImportError("The table contains unsupported null bytes.")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample else ","
    try:
        rows = []
        for row_number, raw in enumerate(csv.reader(io.StringIO(text), delimiter=delimiter), 1):
            values = _trim_row([str(value) for value in raw])
            if not values or not any(str(value).strip() for value in values):
                continue
            if len(values) > MAX_COLUMNS:
                raise RosterImportError(f"Row {row_number} has more than {MAX_COLUMNS} columns.")
            rows.append(
                {
                    "worksheet": worksheet,
                    "row_number": row_number,
                    "raw_values": values,
                }
            )
            if len(rows) > MAX_PREVIEW_ROWS:
                raise RosterImportError(
                    f"A preview may contain at most {MAX_PREVIEW_ROWS} non-empty rows."
                )
    except csv.Error as exc:
        raise RosterImportError("The delimited table could not be parsed.") from exc
    if not rows:
        raise RosterImportError("The table is empty.")
    return rows


def _parse_xlsx(payload: bytes) -> list[dict]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            uncompressed_size = sum(entry.file_size for entry in archive.infolist())
            if uncompressed_size > MAX_UNCOMPRESSED_BYTES:
                raise RosterImportError("The workbook exceeds the 25 MiB uncompressed limit.")
            if any(entry.flag_bits & 0x1 for entry in archive.infolist()):
                raise RosterImportError("Encrypted workbooks are not supported.")
    except zipfile.BadZipFile as exc:
        raise RosterImportError("The uploaded file is not a valid .xlsx workbook.") from exc

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is required in deployed builds
        raise RosterImportError("Excel import support is unavailable.", status_code=503) from exc

    try:
        workbook = load_workbook(
            io.BytesIO(payload),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise RosterImportError("The uploaded file is not a readable .xlsx workbook.") from exc

    parsed = []
    try:
        for worksheet in workbook.worksheets:
            if worksheet.max_column > MAX_COLUMNS:
                raise RosterImportError(
                    f"Worksheet '{worksheet.title}' declares more than {MAX_COLUMNS} columns."
                )
            if worksheet.max_row > MAX_PREVIEW_ROWS + 1:
                raise RosterImportError(
                    f"Worksheet '{worksheet.title}' declares more than {MAX_PREVIEW_ROWS + 1} rows."
                )
            for row_number, cells in enumerate(
                worksheet.iter_rows(
                    max_row=MAX_PREVIEW_ROWS + 1,
                    max_col=MAX_COLUMNS,
                ),
                1,
            ):
                values = _trim_row([_serialized_cell(cell) for cell in cells])
                if not values or not any(value != "" for value in values):
                    continue
                if len(values) > MAX_COLUMNS:
                    raise RosterImportError(
                        f"Worksheet '{worksheet.title}' row {row_number} has more than "
                        f"{MAX_COLUMNS} columns."
                    )
                parsed.append(
                    {
                        "worksheet": worksheet.title,
                        "row_number": row_number,
                        "raw_values": values,
                    }
                )
                if len(parsed) > MAX_PREVIEW_ROWS:
                    raise RosterImportError(
                        f"A workbook preview may contain at most {MAX_PREVIEW_ROWS} non-empty rows."
                    )
    finally:
        workbook.close()
    if not parsed:
        raise RosterImportError("The workbook contains no non-empty rows.")
    return parsed


def parse_roster_source(*, uploaded_file=None, pasted_text=None, requested_source=""):
    if uploaded_file is not None:
        if getattr(uploaded_file, "size", 0) > MAX_UPLOAD_BYTES:
            raise RosterImportError("The uploaded file exceeds the 5 MiB limit.")
        payload = uploaded_file.read(MAX_UPLOAD_BYTES + 1)
        if len(payload) > MAX_UPLOAD_BYTES:
            raise RosterImportError("The uploaded file exceeds the 5 MiB limit.")
        extension = Path(str(uploaded_file.name or "")).suffix.lower()
        if extension == ".xlsx":
            return RosterImportBatch.SourceType.XLSX, "upload.xlsx", _parse_xlsx(payload)
        if extension == ".csv":
            return (
                RosterImportBatch.SourceType.CSV,
                "upload.csv",
                _parse_delimited(payload, worksheet="CSV"),
            )
        if extension == ".xls":
            raise RosterImportError("Legacy .xls files are not supported; save the file as .xlsx.")
        raise RosterImportError("Upload a .csv or .xlsx file.")

    if requested_source and requested_source != RosterImportBatch.SourceType.PASTE:
        raise RosterImportError("sourceType must be 'paste' when no file is uploaded.")
    if not isinstance(pasted_text, str) or not pasted_text.strip():
        raise RosterImportError("pastedText is required when no file is uploaded.")
    payload = pasted_text.encode("utf-8")
    return (
        RosterImportBatch.SourceType.PASTE,
        "pasted table",
        _parse_delimited(payload, worksheet="Pasted data"),
    )
