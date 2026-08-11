import csv
import io
import uuid
import zipfile
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError
from django.test import SimpleTestCase, TestCase
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.test import APIClient

from apps.authn.models import ContactEmail
from apps.authn.tests.helpers import create_member, token_for
from apps.messaging.models import EmailDeliveryRequest, EmailMessageLog
from apps.messaging.services import enqueue_email_job
from apps.scheduling import roster_imports
from apps.scheduling.models import (
    Event,
    EventInvitation,
    EventResultInvalidation,
    FinalMeeting,
    Participant,
    RosterBulkUpdateReceipt,
    RosterImportBatch,
    RosterImportReceipt,
    RosterImportRow,
    UserEvent,
    Weight,
)


class _Cell:
    def __init__(self, value, data_type=""):
        self.value = value
        self.data_type = data_type


class RosterImportParserEdgeTests(SimpleTestCase):
    def test_cell_normalization_and_mapping_helpers_cover_supported_types(self):
        self.assertEqual(roster_imports._trim_row([1, "", None]), [1])
        self.assertEqual(
            roster_imports._serialized_cell(_Cell("SUM(A1)", "f")), {"formula": "SUM(A1)"}
        )
        self.assertEqual(roster_imports._serialized_cell(_Cell(None)), "")
        self.assertEqual(roster_imports._serialized_cell(_Cell(date(2026, 8, 5))), "2026-08-05")
        self.assertEqual(roster_imports._serialized_cell(_Cell(time(9, 30))), "09:30:00")
        self.assertEqual(
            roster_imports._serialized_cell(_Cell(datetime(2026, 8, 5, 9, 30))),
            "2026-08-05T09:30:00",
        )
        self.assertEqual(roster_imports._serialized_cell(_Cell(Decimal("0.25"))), 0.25)
        self.assertEqual(roster_imports._serialized_cell(_Cell(True)), True)
        self.assertEqual(
            roster_imports._serialized_cell(_Cell(object())), str(_Cell(object()).value)
        )

        self.assertEqual(roster_imports._display_cell({"formula": "A1"}), "=A1")
        self.assertEqual(roster_imports._display_cell(None), "")
        self.assertEqual(roster_imports._display_cell(True), "true")
        self.assertEqual(roster_imports._display_cell(False), "false")
        self.assertEqual(roster_imports._canonical_header(" Email_Address "), "email address")
        self.assertEqual(
            roster_imports._auto_mapping(["Participant Name", "E-mail", "Team", "Priority"]),
            {"name": 0, "email": 1, "group": 2, "weight": 3},
        )
        metadata = roster_imports._worksheet_metadata(
            [
                {"worksheet": "Sheet", "row_number": 3, "raw_values": ["A", "B"]},
                {"worksheet": "Sheet", "row_number": 1, "raw_values": ["name"]},
            ]
        )
        self.assertEqual(metadata[0]["defaultHeaderRow"], 1)
        self.assertEqual(metadata[0]["columnCount"], 2)

    def test_mapping_defaults_boolean_and_identity_validation_errors(self):
        headers = ["Name", "Email", "Email"]
        self.assertEqual(roster_imports._mapping_index("1", headers, "email"), 1)
        self.assertEqual(roster_imports._mapping_index(0, headers, "name"), 0)
        for value in (True, object(), -1, 99):
            with self.subTest(value=value), self.assertRaises(roster_imports.RosterImportError):
                roster_imports._mapping_index(value, headers, "name")
        with self.assertRaisesMessage(roster_imports.RosterImportError, "exactly one header"):
            roster_imports._mapping_index("Email", headers, "email")

        self.assertEqual(
            roster_imports._parse_defaults(None),
            {"group": "", "weight": 1.0, "included": True},
        )
        self.assertEqual(
            roster_imports._parse_defaults(
                {"groupName": " Faculty ", "weight": "0.4", "included": 0}
            ),
            {"group": "Faculty", "weight": 0.4, "included": False},
        )
        for defaults, message in [
            ([], "object"),
            ({"group": "x" * 101}, "too long"),
            ({"weight": "bad"}, "between 0 and 1"),
            ({"weight": 2}, "between 0 and 1"),
            ({"included": "sometimes"}, "true or false"),
        ]:
            with (
                self.subTest(defaults=defaults),
                self.assertRaisesMessage(roster_imports.RosterImportError, message),
            ):
                roster_imports._parse_defaults(defaults)

        for value in (True, 1, "YES", "included"):
            self.assertTrue(roster_imports._parse_included(value))
        for value in (False, 0, "NO", "excluded"):
            self.assertFalse(roster_imports._parse_included(value))
        with self.assertRaisesMessage(roster_imports.RosterImportError, "true or false"):
            roster_imports._parse_included("maybe")

        self.assertEqual(
            roster_imports._validate_identity_fields("", "", ""),
            ["name is required.", "email is required."],
        )
        errors = roster_imports._validate_identity_fields(
            "n" * 101,
            "e" * 255,
            "g" * 101,
        )
        self.assertIn("name is too long (max 100).", errors)
        self.assertIn("email is too long (max 254).", errors)
        self.assertIn("group is too long (max 100).", errors)
        self.assertEqual(
            roster_imports._validate_identity_fields("Name", "not-an-email", ""),
            ["email is invalid."],
        )

    def test_delimited_parser_rejects_malformed_and_bounded_input(self):
        with (
            patch.object(roster_imports, "MAX_UNCOMPRESSED_BYTES", 2),
            self.assertRaisesMessage(roster_imports.RosterImportError, "25 MiB"),
        ):
            roster_imports._parse_delimited(b"abc", worksheet="CSV")
        with self.assertRaisesMessage(roster_imports.RosterImportError, "UTF-8"):
            roster_imports._parse_delimited(b"\xff", worksheet="CSV")
        with self.assertRaisesMessage(roster_imports.RosterImportError, "null bytes"):
            roster_imports._parse_delimited(b"name\x00,email", worksheet="CSV")
        with self.assertRaisesMessage(roster_imports.RosterImportError, "empty"):
            roster_imports._parse_delimited(b"\n,\n", worksheet="CSV")

        with patch("apps.scheduling.roster_imports.csv.Sniffer.sniff", side_effect=csv.Error):
            parsed = roster_imports._parse_delimited(
                b"name\temail\nAda\tada@example.com", worksheet="P"
            )
        self.assertEqual(parsed[1]["raw_values"], ["Ada", "ada@example.com"])

        with (
            patch.object(roster_imports, "MAX_COLUMNS", 1),
            self.assertRaisesMessage(roster_imports.RosterImportError, "more than 1 columns"),
        ):
            roster_imports._parse_delimited(b"a,b", worksheet="CSV")
        with (
            patch.object(roster_imports, "MAX_PREVIEW_ROWS", 1),
            self.assertRaisesMessage(roster_imports.RosterImportError, "at most 1"),
        ):
            roster_imports._parse_delimited(b"a\nb", worksheet="CSV")

        def broken_reader(*_args, **_kwargs):
            yield ["name"]
            raise csv.Error("broken quoting")

        with (
            patch("apps.scheduling.roster_imports.csv.reader", broken_reader),
            self.assertRaisesMessage(roster_imports.RosterImportError, "could not be parsed"),
        ):
            roster_imports._parse_delimited(b"name", worksheet="CSV")

    def test_xlsx_parser_rejects_invalid_encrypted_unreadable_empty_and_bounds(self):
        with self.assertRaisesMessage(roster_imports.RosterImportError, "valid .xlsx"):
            roster_imports._parse_xlsx(b"not a zip")

        fake_archive = MagicMock()
        fake_archive.__enter__.return_value = fake_archive
        fake_archive.__exit__.return_value = False
        fake_archive.infolist.return_value = [SimpleNamespace(file_size=1, flag_bits=1)]
        with patch("apps.scheduling.roster_imports.zipfile.ZipFile", return_value=fake_archive):
            with self.assertRaisesMessage(roster_imports.RosterImportError, "Encrypted"):
                roster_imports._parse_xlsx(b"zip")

        valid_zip = io.BytesIO()
        with zipfile.ZipFile(valid_zip, "w") as archive:
            archive.writestr("placeholder", "x")
        with patch("openpyxl.load_workbook", side_effect=ValueError("bad workbook")):
            with self.assertRaisesMessage(roster_imports.RosterImportError, "readable"):
                roster_imports._parse_xlsx(valid_zip.getvalue())

        blank = Workbook()
        blank_output = io.BytesIO()
        blank.save(blank_output)
        with self.assertRaisesMessage(roster_imports.RosterImportError, "no non-empty"):
            roster_imports._parse_xlsx(blank_output.getvalue())

        workbook = Workbook()
        workbook.active.append(["name", "email"])
        workbook.active.append(["Ada", "ada@example.com"])
        output = io.BytesIO()
        workbook.save(output)
        with (
            patch.object(roster_imports, "MAX_COLUMNS", 1),
            self.assertRaisesMessage(roster_imports.RosterImportError, "more than 1 columns"),
        ):
            roster_imports._parse_xlsx(output.getvalue())
        with (
            patch.object(roster_imports, "MAX_PREVIEW_ROWS", 1),
            self.assertRaisesMessage(roster_imports.RosterImportError, "at most 1"),
        ):
            roster_imports._parse_xlsx(output.getvalue())

        declared_large = Workbook()
        declared_large.active.append(["name", "email"])
        declared_large.active.append(["One", "one@example.com"])
        declared_large.active.append(["Two", "two@example.com"])
        declared_large_output = io.BytesIO()
        declared_large.save(declared_large_output)
        with (
            patch.object(roster_imports, "MAX_PREVIEW_ROWS", 1),
            self.assertRaisesMessage(roster_imports.RosterImportError, "declares more than 2"),
        ):
            roster_imports._parse_xlsx(declared_large_output.getvalue())

        fake_worksheet = MagicMock(max_column=1, max_row=1, title="Mock")
        fake_worksheet.iter_rows.return_value = [[_Cell("name"), _Cell("email")]]
        fake_workbook = MagicMock(worksheets=[fake_worksheet])
        with (
            patch.object(roster_imports, "MAX_COLUMNS", 1),
            patch("openpyxl.load_workbook", return_value=fake_workbook),
            self.assertRaisesMessage(roster_imports.RosterImportError, "row 1 has more than 1"),
        ):
            roster_imports._parse_xlsx(valid_zip.getvalue())
        fake_workbook.close.assert_called_once()

        blank_then_value = Workbook()
        blank_then_value.active["A2"] = "name"
        blank_then_value_output = io.BytesIO()
        blank_then_value.save(blank_then_value_output)
        self.assertEqual(
            roster_imports._parse_xlsx(blank_then_value_output.getvalue())[0]["row_number"],
            2,
        )

    def test_source_selection_and_upload_validation(self):
        oversized = SimpleNamespace(size=3, name="roster.csv", read=lambda _size: b"abc")
        with (
            patch.object(roster_imports, "MAX_UPLOAD_BYTES", 2),
            self.assertRaisesMessage(roster_imports.RosterImportError, "5 MiB"),
        ):
            roster_imports.parse_roster_source(uploaded_file=oversized)

        deceptive = SimpleNamespace(size=0, name="roster.csv", read=lambda _size: b"abc")
        with (
            patch.object(roster_imports, "MAX_UPLOAD_BYTES", 2),
            self.assertRaisesMessage(roster_imports.RosterImportError, "5 MiB"),
        ):
            roster_imports.parse_roster_source(uploaded_file=deceptive)

        csv_upload = SimpleUploadedFile("people.csv", b"name,email\nA,a@example.com")
        source, label, rows = roster_imports.parse_roster_source(uploaded_file=csv_upload)
        self.assertEqual((source, label, rows[0]["worksheet"]), ("csv", "upload.csv", "CSV"))

        for filename, message in [("legacy.xls", "Legacy .xls"), ("people.txt", "csv or .xlsx")]:
            with (
                self.subTest(filename=filename),
                self.assertRaisesMessage(roster_imports.RosterImportError, message),
            ):
                roster_imports.parse_roster_source(
                    uploaded_file=SimpleUploadedFile(filename, b"anything")
                )
        with self.assertRaisesMessage(roster_imports.RosterImportError, "sourceType"):
            roster_imports.parse_roster_source(pasted_text="a", requested_source="csv")
        with self.assertRaisesMessage(roster_imports.RosterImportError, "pastedText"):
            roster_imports.parse_roster_source(pasted_text=" ")

    def test_row_normalization_and_duplicate_reconciliation(self):
        row = RosterImportRow(
            raw_values=[{"formula": "NAME"}, "=EMAIL", {"formula": "GROUP"}, "2", "maybe"]
        )
        roster_imports._normalize_row(
            row,
            {"name": 0, "email": 1, "group": 2, "weight": 3, "included": 4},
            {"group": "", "weight": 0.5, "included": False},
        )
        self.assertTrue(row.selected)
        self.assertIn("name cannot contain a formula.", row.validation_errors)
        self.assertIn("email cannot contain a formula.", row.validation_errors)
        self.assertIn("group cannot contain a formula.", row.validation_errors)
        self.assertIn("weight must be between 0 and 1.", row.validation_errors)
        self.assertIn("included must be true or false.", row.validation_errors)

        missing = RosterImportRow(raw_values=[])
        roster_imports._normalize_row(missing, {}, {"group": "", "weight": 1, "included": True})
        self.assertIn("Map a name column.", missing.validation_errors)
        self.assertIn("Map an email column.", missing.validation_errors)

        first = RosterImportRow(
            name="A",
            email="same@example.com",
            selected=True,
            validation_errors=[],
            duplicate_status=RosterImportRow.DuplicateStatus.CONFLICT,
        )
        second = RosterImportRow(
            name="B",
            email="same@example.com",
            selected=False,
            validation_errors=["Conflicting duplicate email."],
            duplicate_status=RosterImportRow.DuplicateStatus.CONFLICT,
        )
        roster_imports._apply_duplicate_rules([first, second])
        self.assertEqual(first.duplicate_status, RosterImportRow.DuplicateStatus.UNIQUE)
        self.assertEqual(second.duplicate_status, RosterImportRow.DuplicateStatus.UNIQUE)
        self.assertEqual(second.validation_errors, [])
        self.assertEqual(roster_imports._active_rows(SimpleNamespace(selected_worksheet="")), [])

        formula_options = RosterImportRow(
            raw_values=["Name", "name@example.com", {"formula": "WEIGHT"}, "=INCLUDED"]
        )
        roster_imports._normalize_row(
            formula_options,
            {"name": 0, "email": 1, "weight": 2, "included": 3},
            {"group": "", "weight": 1, "included": True},
        )
        self.assertIn("weight cannot contain a formula.", formula_options.validation_errors)
        self.assertIn("included cannot contain a formula.", formula_options.validation_errors)

        tail = RosterImportRow(name="Tail", email="tail@example.com", selected=True)
        second.duplicate_status = RosterImportRow.DuplicateStatus.UNIQUE
        roster_imports._apply_duplicate_rules([second, tail])
        self.assertEqual(second.duplicate_status, RosterImportRow.DuplicateStatus.UNIQUE)


class RosterImportDatabaseEdgeTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.organizer = create_member("roster-edge-owner@example.com", "Roster", "Owner")
        self.outsider = create_member("roster-edge-outsider@example.com", "Outside", "User")
        self.event = Event.objects.create(
            code="ROSTEDGE",
            name="Roster edge cases",
            organizer=self.organizer,
            days=[1],
            start_minutes=9 * 60,
            end_minutes=10 * 60,
            slot_minutes=30,
        )
        UserEvent.objects.create(member=self.organizer, event=self.event, role="organizer")
        self.authenticate(self.organizer)

    def authenticate(self, member):
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token_for(member)}")

    def preview(self, content="name,email\nAda,ada-edge@example.com"):
        response = self.client.post(
            f"/events/roster-imports?code={self.event.code}",
            {"sourceType": "paste", "pastedText": content},
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        return RosterImportBatch.objects.get(pk=response.data["import"]["id"])

    def commit(self, batch, **overrides):
        payload = {"mode": "merge", "idempotencyKey": str(uuid.uuid4()), **overrides}
        return self.client.post(
            f"/events/roster-imports/{batch.pk}/commit?code={self.event.code}",
            payload,
            format="json",
        )

    def test_preview_update_validation_and_no_selected_worksheet(self):
        batch = self.preview()
        batch.selected_worksheet = ""
        batch.save(update_fields=["selected_worksheet", "updated_at"])
        roster_imports.normalize_import_batch(batch)
        batch.refresh_from_db()
        self.assertEqual(batch.summary["total"], 0)
        with self.assertRaisesMessage(roster_imports.RosterImportError, "Select a worksheet"):
            roster_imports.update_roster_import(batch=batch, data={"headerRow": 1})

        batch = self.preview()
        cases = [
            ({"worksheet": "missing"}, "worksheet was not found"),
            ({"headerRow": True}, "positive integer"),
            ({"headerRow": 0}, "positive integer"),
            ({"headerRow": 999}, "does not identify"),
            ({"columnMapping": []}, "must be an object"),
            ({"columnMapping": {"unknown": 0}}, "Unknown mapped field"),
            ({"columnMapping": {"name": 0}}, "include name and email"),
            ({"columnMapping": {"name": 0, "email": 0}}, "different column"),
            ({"columnMapping": {"name": True, "email": 1}}, "identify a column"),
            ({"defaults": []}, "defaults must be an object"),
        ]
        for data, message in cases:
            with (
                self.subTest(data=data),
                self.assertRaisesMessage(roster_imports.RosterImportError, message),
            ):
                roster_imports.update_roster_import(batch=batch, data=data)

        refreshed = roster_imports.update_roster_import(
            batch=batch,
            data={"worksheet": "Pasted data", "headerRow": 1},
        )
        self.assertEqual(refreshed.column_mapping, {"name": 0, "email": 1})

        header_only = self.preview("name,email")
        self.assertEqual(header_only.summary["total"], 0)

        ghost = self.preview("name,email\nGhost,ghost-edge@example.com")
        ghost.worksheets = [{"name": "Ghost", "headers": ["name", "email"]}]
        ghost.save(update_fields=["worksheets", "updated_at"])
        with self.assertRaisesMessage(roster_imports.RosterImportError, "no non-empty rows"):
            roster_imports.update_roster_import(batch=ghost, data={"worksheet": "Ghost"})

        payload_batch = self.preview("name,email\nPayload,payload-edge@example.com")
        payload_batch.rows.filter(row_number=1).delete()
        payload = roster_imports.roster_import_payload(payload_batch)
        self.assertEqual(payload["headers"], ["name", "email"])

    def test_row_update_validation_and_every_editable_field(self):
        batch = self.preview()
        row = batch.rows.get(row_number=2)
        for updates, message in [
            ({"rowUpdates": {}}, "must be an array"),
            ({"rowUpdates": [{"id": str(row.pk)}, {"id": str(row.pk)}]}, "unique id"),
            ({"rowUpdates": [{"id": str(uuid.uuid4())}]}, "does not belong"),
            ({"rowUpdates": [{"id": str(row.pk), "weight": "bad"}]}, "between 0 and 1"),
            ({"rowUpdates": [{"id": str(row.pk), "weight": -1}]}, "between 0 and 1"),
            ({"rowUpdates": [{"id": str(row.pk), "included": "maybe"}]}, "true or false"),
            ({"rowUpdates": [{"id": str(row.pk), "selected": "maybe"}]}, "true or false"),
        ]:
            with (
                self.subTest(updates=updates),
                self.assertRaisesMessage(roster_imports.RosterImportError, message),
            ):
                roster_imports.update_roster_import(batch=batch, data=updates)

        with (
            patch.object(roster_imports, "MAX_ROSTER_ROWS", 1),
            self.assertRaisesMessage(roster_imports.RosterImportError, "at most 1"),
        ):
            roster_imports.update_roster_import(
                batch=batch,
                data={"rowUpdates": [{"id": str(row.pk)}, {"id": str(uuid.uuid4())}]},
            )

        updated = roster_imports.update_roster_import(
            batch=batch,
            data={
                "rowUpdates": [
                    {
                        "id": str(row.pk),
                        "name": " Grace ",
                        "email": " GRACE-EDGE@EXAMPLE.COM ",
                        "groupName": " Faculty ",
                        "weight": "0.3",
                        "included": "false",
                        "selected": "true",
                    }
                ]
            },
        )
        edited = updated.rows.get(pk=row.pk)
        self.assertEqual(edited.name, "Grace")
        self.assertEqual(edited.email, "grace-edge@example.com")
        self.assertEqual(edited.group_name, "Faculty")
        self.assertEqual(edited.weight, 0.3)
        self.assertFalse(edited.included)

        two_rows = self.preview(
            "name,email\nOne,one-update-cap@example.com\nTwo,two-update-cap@example.com"
        )
        one_row = two_rows.rows.get(row_number=2)
        with (
            patch.object(roster_imports, "MAX_ROSTER_ROWS", 1),
            self.assertRaisesMessage(roster_imports.RosterImportError, "at most 1 valid"),
        ):
            roster_imports.update_roster_import(
                batch=two_rows,
                data={"rowUpdates": [{"id": str(one_row.pk), "name": "Still valid"}]},
            )

    def test_expired_committed_and_repeated_cancel_contracts(self):
        expired = self.preview()
        expired.expires_at = timezone.now() - timedelta(seconds=1)
        expired.save(update_fields=["expires_at", "updated_at"])
        with self.assertRaisesMessage(roster_imports.RosterImportError, "expired") as raised:
            roster_imports.update_roster_import(batch=expired, data={})
        self.assertEqual(raised.exception.status_code, 410)
        expired.status = RosterImportBatch.Status.EXPIRED
        expired.save(update_fields=["status", "updated_at"])
        self.assertTrue(roster_imports.expire_roster_import_preview(expired))

        canceled = self.preview("name,email\nCancel,cancel-edge@example.com")
        roster_imports.cancel_roster_import(canceled)
        repeated = roster_imports.cancel_roster_import(canceled)
        self.assertEqual(repeated.status, RosterImportBatch.Status.CANCELED)

        committed = self.preview("name,email\nCommit,commit-edge@example.com")
        self.assertEqual(self.commit(committed).status_code, 201)
        with self.assertRaisesMessage(roster_imports.RosterImportError, "cannot be canceled"):
            roster_imports.cancel_roster_import(committed)
        with self.assertRaisesMessage(roster_imports.RosterImportError, "can no longer be changed"):
            roster_imports.update_roster_import(batch=committed, data={})

    def test_expiry_cleanup_tolerates_a_concurrent_preview_removal(self):
        batch = self.preview("name,email\nRace,race-expiry-edge@example.com")
        batch.expires_at = timezone.now() - timedelta(seconds=1)
        batch.save(update_fields=["expires_at", "updated_at"])
        locked_query = MagicMock()
        locked_query.filter.return_value.first.return_value = None
        with patch.object(
            RosterImportBatch.objects,
            "select_for_update",
            return_value=locked_query,
        ):
            self.assertEqual(roster_imports.expire_stale_roster_imports(), 1)
        batch.refresh_from_db()
        self.assertEqual(batch.status, RosterImportBatch.Status.PREVIEW)

    def test_rebuild_without_temporary_challenges_uses_the_empty_cleanup_path(self):
        self.assertFalse(self.event.result_invalidations.exists())
        roster_imports._rebuild_event_roster(self.event, timezone.now())
        self.assertEqual(self.event.status, Event.Status.DRAFT)
        self.assertEqual(self.event.version, 2)

    def test_member_resolution_rejects_unsafe_identities_and_adopts_orphan(self):
        inactive = create_member("inactive-edge@example.com", is_active=False)
        unverified = create_member("unverified-edge@example.com", contact_verified=False)
        for email, expected in [
            (inactive.email, "inactive account"),
            (unverified.email, "unverified full account"),
        ]:
            row = RosterImportRow(name="Unsafe", email=email)
            with (
                self.subTest(email=email),
                self.assertRaisesMessage(roster_imports.RosterImportError, expected),
            ):
                roster_imports._resolve_members([row])

        orphan = ContactEmail.objects.create(
            member=None,
            email_address="orphan-edge@example.com",
            email_type="other",
            verified=True,
        )
        resolved = roster_imports._resolve_members(
            [RosterImportRow(name="Orphan", email="orphan-edge@example.com")]
        )
        orphan.refresh_from_db()
        self.assertEqual(orphan.member, resolved["orphan-edge@example.com"])
        self.assertEqual(orphan.email_type, "primary")
        self.assertFalse(orphan.verified)

        shared = create_member("shared-primary-edge@example.com")
        ContactEmail.objects.create(
            member=shared,
            email_address="shared-other-edge@example.com",
            email_type="secondary",
            verified=True,
        )
        with self.assertRaisesMessage(roster_imports.RosterImportError, "same account"):
            roster_imports._resolve_members(
                [
                    RosterImportRow(name="One", email="shared-primary-edge@example.com"),
                    RosterImportRow(name="Two", email="shared-other-edge@example.com"),
                ]
            )

        locked_query = MagicMock()
        locked_query.filter.return_value.order_by.return_value = []
        with (
            patch.object(
                type(shared).objects,
                "select_for_update",
                return_value=locked_query,
            ),
            self.assertRaisesMessage(roster_imports.RosterImportError, "account changed"),
        ):
            roster_imports._resolve_members(
                [
                    RosterImportRow(
                        name="Concurrent",
                        email="shared-primary-edge@example.com",
                    )
                ]
            )

    def test_merge_updates_existing_identity_invitation_weight_and_enforces_capacity(self):
        member = create_member("existing-roster-edge@example.com")
        participant = Participant.objects.create(
            event=self.event,
            member=member,
            participant_name="Old",
            group_name="Old group",
            hidden=True,
            availability_inperson=[1, 0],
            availability_virtual=[0, 1],
        )
        invitation = EventInvitation.objects.create(
            event=self.event,
            email=member.email,
            member=None,
            invited_by=None,
        )
        Weight.objects.create(event=self.event, participant=participant, weight=1, included=True)
        batch = self.preview(
            "name,email,group,weight,included\n"
            "Renamed,existing-roster-edge@example.com,New group,0.2,false"
        )
        committed = self.commit(batch)
        self.assertEqual(committed.status_code, 201, committed.data)
        self.assertEqual(committed.data["event"]["code"], self.event.code)
        self.assertEqual(committed.data["event"]["status"], Event.Status.DRAFT)
        participant.refresh_from_db()
        invitation.refresh_from_db()
        self.assertEqual(participant.participant_name, "Renamed")
        self.assertFalse(participant.hidden)
        self.assertEqual(invitation.member, member)
        self.assertEqual(invitation.invited_by, self.organizer)
        weight = Weight.objects.get(participant=participant)
        self.assertEqual(weight.weight, 0.2)
        self.assertFalse(weight.included)

        unchanged = self.preview(
            "name,email,group,weight,included\n"
            "Renamed,existing-roster-edge@example.com,New group,0.2,false"
        )
        repeated_merge = self.commit(unchanged)
        self.assertEqual(repeated_merge.status_code, 201, repeated_merge.data)
        self.assertEqual(repeated_merge.data["receipt"]["updatedCount"], 1)

        over_capacity = self.preview("name,email\nExtra,extra-capacity-edge@example.com")
        with patch.object(roster_imports, "MAX_ROSTER_ROWS", 1):
            response = self.commit(over_capacity)
        self.assertEqual(response.status_code, 409)

        EventInvitation.objects.create(
            event=self.event,
            email="invitation-only-capacity@example.com",
            invited_by=self.organizer,
        )
        invitation_capacity = self.preview(
            "name,email\nInvitation Cap,new-invitation-capacity@example.com"
        )
        with patch.object(roster_imports, "MAX_ROSTER_ROWS", 2):
            invitation_capacity_response = self.commit(invitation_capacity)
        self.assertEqual(invitation_capacity_response.status_code, 409)
        self.assertIn("invitation recipients", invitation_capacity_response.data["error"])

        oversized_batch = self.preview(
            "name,email\nFirst,first-commit-cap@example.com\nSecond,second-commit-cap@example.com"
        )
        with patch.object(roster_imports, "MAX_ROSTER_ROWS", 1):
            oversized_response = self.commit(oversized_batch)
        self.assertEqual(oversized_response.status_code, 400)
        self.assertIn("at most 1", oversized_response.data["error"])

    def test_commit_rejects_invalid_state_actor_rows_and_integrity_race(self):
        batch = self.preview()
        for payload, message in [
            ({"mode": "unknown", "idempotencyKey": str(uuid.uuid4())}, "mode"),
            ({"mode": "merge", "idempotencyKey": "bad"}, "UUID"),
        ]:
            with (
                self.subTest(payload=payload),
                self.assertRaisesMessage(roster_imports.RosterImportError, message),
            ):
                roster_imports.commit_roster_import(
                    event=self.event,
                    batch_id=batch.pk,
                    organizer=self.organizer,
                    data=payload,
                )

        with self.assertRaisesMessage(roster_imports.RosterImportError, "not found"):
            roster_imports.commit_roster_import(
                event=self.event,
                batch_id=uuid.uuid4(),
                organizer=self.organizer,
                data={"mode": "merge", "idempotencyKey": str(uuid.uuid4())},
            )
        with self.assertRaisesMessage(roster_imports.RosterImportError, "Only the organizer"):
            roster_imports.commit_roster_import(
                event=self.event,
                batch_id=batch.pk,
                organizer=self.outsider,
                data={"mode": "merge", "idempotencyKey": str(uuid.uuid4())},
            )

        self.event.status = Event.Status.FINALIZED
        self.event.save(update_fields=["status", "updated_at"])
        with self.assertRaisesMessage(roster_imports.RosterImportError, "Reopen"):
            roster_imports.commit_roster_import(
                event=self.event,
                batch_id=batch.pk,
                organizer=self.organizer,
                data={"mode": "merge", "idempotencyKey": str(uuid.uuid4())},
            )
        self.event.status = Event.Status.DRAFT
        self.event.save(update_fields=["status", "updated_at"])

        final_meeting = FinalMeeting.objects.create(
            event=self.event,
            starts_at=timezone.now() + timedelta(days=1),
            ends_at=timezone.now() + timedelta(days=1, minutes=30),
            timezone="UTC",
            channel="inperson",
            location="Room",
            calendar_uid=f"final-{self.event.event_id}@releviz",
            confirmed_by=self.organizer,
            confirmed_at=timezone.now(),
        )
        with self.assertRaisesMessage(roster_imports.RosterImportError, "confirmed meeting"):
            roster_imports.commit_roster_import(
                event=self.event,
                batch_id=batch.pk,
                organizer=self.organizer,
                data={"mode": "merge", "idempotencyKey": str(uuid.uuid4())},
            )
        final_meeting.delete()

        row = batch.rows.get(row_number=2)
        row.selected = False
        row.save(update_fields=["selected", "updated_at"])
        with self.assertRaisesMessage(roster_imports.RosterImportError, "Select at least one"):
            roster_imports.commit_roster_import(
                event=self.event,
                batch_id=batch.pk,
                organizer=self.organizer,
                data={"mode": "merge", "idempotencyKey": str(uuid.uuid4())},
            )
        row.selected = True
        row.validation_errors = ["invalid"]
        row.save(update_fields=["selected", "validation_errors", "updated_at"])
        with self.assertRaisesMessage(roster_imports.RosterImportError, "Resolve or deselect"):
            roster_imports.commit_roster_import(
                event=self.event,
                batch_id=batch.pk,
                organizer=self.organizer,
                data={"mode": "merge", "idempotencyKey": str(uuid.uuid4())},
            )
        row.validation_errors = []
        row.save(update_fields=["validation_errors", "updated_at"])

        with patch("apps.scheduling.roster_imports._write_roster", side_effect=IntegrityError):
            with self.assertRaisesMessage(roster_imports.RosterImportError, "changed concurrently"):
                roster_imports.commit_roster_import(
                    event=self.event,
                    batch_id=batch.pk,
                    organizer=self.organizer,
                    data={"mode": "merge", "idempotencyKey": str(uuid.uuid4())},
                )

    def test_import_endpoints_cover_errors_pagination_and_committed_cancel(self):
        self.assertEqual(
            self.client.post("/events/roster-imports", {}, format="json").status_code, 400
        )
        self.assertEqual(
            self.client.post("/events/roster-imports?code=MISSING", {}, format="json").status_code,
            404,
        )
        self.event.status = Event.Status.ARCHIVED
        self.event.save(update_fields=["status", "updated_at"])
        blocked = self.client.post(
            f"/events/roster-imports?code={self.event.code}",
            {"sourceType": "paste", "pastedText": "name,email\nA,a2@example.com"},
            format="json",
        )
        self.assertEqual(blocked.status_code, 409)
        self.event.status = Event.Status.DRAFT
        self.event.save(update_fields=["status", "updated_at"])

        missing_id = uuid.uuid4()
        self.assertEqual(
            self.client.put(
                f"/events/roster-imports/{missing_id}?code={self.event.code}", {}, format="json"
            ).status_code,
            404,
        )
        self.assertEqual(
            self.client.delete(
                f"/events/roster-imports/{missing_id}?code={self.event.code}"
            ).status_code,
            404,
        )
        self.assertEqual(
            self.client.get(
                f"/events/roster-imports/{missing_id}/rows?code={self.event.code}"
            ).status_code,
            404,
        )

        batch = self.preview("name,email\nA,a3@example.com\nB,b3@example.com")
        expired_batch = self.preview("name,email\nExpired,expired-delete@example.com")
        expired_batch.expires_at = timezone.now() - timedelta(seconds=1)
        expired_batch.save(update_fields=["expires_at", "updated_at"])
        expired_delete = self.client.delete(
            f"/events/roster-imports/{expired_batch.pk}?code={self.event.code}"
        )
        self.assertEqual(expired_delete.status_code, 410)
        invalid_update = self.client.put(
            f"/events/roster-imports/{batch.pk}?code={self.event.code}",
            {"headerRow": 0},
            format="json",
        )
        self.assertEqual(invalid_update.status_code, 400)
        invalid_page = self.client.get(
            f"/events/roster-imports/{batch.pk}/rows?code={self.event.code}&page=bad"
        )
        self.assertEqual(invalid_page.status_code, 400)
        invalid_size = self.client.get(
            f"/events/roster-imports/{batch.pk}/rows?code={self.event.code}&pageSize=101"
        )
        self.assertEqual(invalid_size.status_code, 400)
        page = self.client.get(
            f"/events/roster-imports/{batch.pk}/rows?code={self.event.code}&pageSize=1&page=2"
        )
        self.assertEqual(page.status_code, 200)
        self.assertEqual(page.data["pagination"]["pages"], 2)
        self.assertEqual(len(page.data["rows"]), 1)

        self.assertEqual(self.commit(batch).status_code, 201)
        cannot_cancel = self.client.delete(
            f"/events/roster-imports/{batch.pk}?code={self.event.code}"
        )
        self.assertEqual(cannot_cancel.status_code, 409)

    def test_roster_endpoints_consistently_reject_non_organizers(self):
        batch = self.preview("name,email\nDenied,denied-edge@example.com")
        participant_id = uuid.uuid4()
        self.authenticate(self.outsider)
        requests = [
            self.client.put(
                f"/events/roster-imports/{batch.pk}?code={self.event.code}",
                {},
                format="json",
            ),
            self.client.delete(f"/events/roster-imports/{batch.pk}?code={self.event.code}"),
            self.client.get(f"/events/roster-imports/{batch.pk}/rows?code={self.event.code}"),
            self.client.post(
                f"/events/roster-imports/{batch.pk}/commit?code={self.event.code}",
                {},
                format="json",
            ),
            self.client.get(f"/events/roster/{participant_id}/schedule?code={self.event.code}"),
            self.client.patch(
                f"/events/roster/{participant_id}?code={self.event.code}",
                {},
                format="json",
            ),
            self.client.patch(
                f"/events/roster/bulk?code={self.event.code}",
                {},
                format="json",
            ),
        ]
        self.assertEqual([response.status_code for response in requests], [403] * len(requests))

    def test_roster_filters_schedule_and_participant_patch_edge_contracts(self):
        batch = self.preview(
            "name,email,group,included\n"
            "Ada,ada-filter-edge@example.com,Faculty,true\n"
            "Grace,grace-filter-edge@example.com,,false"
        )
        self.assertEqual(self.commit(batch).status_code, 201)
        ada = self.event.participants.get(participant_name="Ada")
        grace = self.event.participants.get(participant_name="Grace")
        ada.submitted = True
        ada.save(update_fields=["submitted", "updated_at"])
        invitation = self.event.invitations.get(member=grace.member)
        invitation.first_sent_at = timezone.now()
        invitation.save(update_fields=["first_sent_at", "updated_at"])
        delivery_job, _created = enqueue_email_job(
            idempotency_key="roster-latest-delivery",
            message_type=EmailMessageLog.MessageType.INVITATION,
            recipient=grace.member.email,
            subject="Roster delivery",
            body="Invitation",
            message_id="<roster-latest-delivery@releviz.local>",
            event=self.event,
            invitation=invitation,
        )
        delivery_request = EmailDeliveryRequest.objects.create(
            event=self.event,
            requested_by=self.organizer,
            operation=EmailDeliveryRequest.Operation.INVITATION,
            idempotency_key=uuid.uuid4(),
            request_fingerprint="d" * 64,
            recipient_count=1,
            created_job_count=1,
        )
        delivery_request.jobs.add(delivery_job)

        for query, total in [
            ("search=ada-filter", 1),
            ("group=Faculty", 1),
            ("group=__ungrouped__", 1),
            ("submitted=yes", 1),
            ("submitted=no", 1),
            ("included=true", 1),
            ("included=false", 1),
            ("invitationStatus=submitted", 1),
            ("invitationStatus=invited", 1),
            ("accountAccess=temporary", 2),
        ]:
            with self.subTest(query=query):
                response = self.client.get(f"/events/roster?code={self.event.code}&{query}")
                self.assertEqual(response.status_code, 200, response.data)
                self.assertEqual(response.data["pagination"]["total"], total)
                self.assertEqual(
                    response.data["latestDeliveryRequest"]["delivery"]["pending"],
                    1,
                )
        for query in [
            "submitted=maybe",
            "included=maybe",
            "invitationStatus=bad",
            "accountAccess=bad",
            "page=0",
        ]:
            with self.subTest(query=query):
                self.assertEqual(
                    self.client.get(f"/events/roster?code={self.event.code}&{query}").status_code,
                    400,
                )

        missing = uuid.uuid4()
        self.assertEqual(
            self.client.get(
                f"/events/roster/{missing}/schedule?code={self.event.code}"
            ).status_code,
            404,
        )
        self.assertEqual(
            self.client.patch(
                f"/events/roster/{missing}?code={self.event.code}",
                {"expectedVersion": 1},
                format="json",
            ).status_code,
            404,
        )
        self.assertEqual(
            self.client.patch(
                f"/events/roster/{ada.pk}?code={self.event.code}", {}, format="json"
            ).status_code,
            428,
        )

        for payload in [
            {"expectedVersion": ada.version, "name": ""},
            {"expectedVersion": ada.version, "name": "x" * 101},
            {"expectedVersion": ada.version, "group": "x" * 101},
            {"expectedVersion": ada.version, "weight": "bad"},
            {"expectedVersion": ada.version, "weight": 2},
        ]:
            with self.subTest(payload=payload):
                response = self.client.patch(
                    f"/events/roster/{ada.pk}?code={self.event.code}", payload, format="json"
                )
                self.assertEqual(response.status_code, 400, response.data)

        unchanged = self.client.patch(
            f"/events/roster/{ada.pk}?code={self.event.code}",
            {
                "expectedVersion": ada.version,
                "name": "Ada",
                "group": "Faculty",
                "weight": 1,
                "included": True,
            },
            format="json",
        )
        self.assertEqual(unchanged.status_code, 200, unchanged.data)
        self.assertEqual(unchanged.data["participant"]["version"], ada.version)

        direct_member = create_member("direct-weight-edge@example.com")
        direct = Participant.objects.create(
            event=self.event,
            member=direct_member,
            participant_name="Direct",
            availability_inperson=[0, 0],
            availability_virtual=[0, 0],
        )
        created_weight = self.client.patch(
            f"/events/roster/{direct.pk}?code={self.event.code}",
            {
                "expectedVersion": direct.version,
                "name": "Direct renamed",
                "groupName": "",
                "included": "false",
            },
            format="json",
        )
        self.assertEqual(created_weight.status_code, 200, created_weight.data)
        self.assertFalse(Weight.objects.get(participant=direct).included)

        self.event.status = Event.Status.FINALIZED
        self.event.save(update_fields=["status", "updated_at"])
        blocked = self.client.patch(
            f"/events/roster/{direct.pk}?code={self.event.code}",
            {"expectedVersion": direct.version},
            format="json",
        )
        self.assertEqual(blocked.status_code, 409)

    def test_bulk_selectors_validation_new_and_existing_weights(self):
        batch = self.preview(
            "name,email,group\nOne,one-bulk-edge@example.com,A\nTwo,two-bulk-edge@example.com,"
        )
        self.assertEqual(self.commit(batch).status_code, 201)
        participants = list(self.event.participants.order_by("participant_name"))

        missing_key = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {"group": "A", "updates": {"group": "X"}},
            format="json",
        )
        self.assertEqual(missing_key.status_code, 400)
        self.assertIn("idempotencyKey", missing_key.data["error"])

        cases = [
            ({}, "updates"),
            ({"participantIds": [], "updates": {"group": "X"}}, "non-empty"),
            ({"updates": {"group": "X"}}, "Choose"),
            ({"group": "A", "updates": {"unknown": 1}}, "Unknown"),
            ({"filter": [], "updates": {"group": "X"}}, "filter must be an object"),
            ({"group": "A", "updates": {"group": "x" * 101}}, "too long"),
        ]
        for payload, message in cases:
            with self.subTest(payload=payload):
                response = self.client.patch(
                    f"/events/roster/bulk?code={self.event.code}",
                    {**payload, "idempotencyKey": str(uuid.uuid4())},
                    format="json",
                )
                self.assertEqual(response.status_code, 400, response.data)
                self.assertIn(message, response.data["error"])

        unknown_request = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                "group": "A",
                "updates": {"group": "X"},
                "unexpected": True,
                "idempotencyKey": str(uuid.uuid4()),
            },
            format="json",
        )
        self.assertEqual(unknown_request.status_code, 400)
        self.assertIn("Unknown bulk request field", unknown_request.data["error"])

        invalid_uuid = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                "participantIds": ["bad"],
                "updates": {"group": "X"},
                "idempotencyKey": str(uuid.uuid4()),
            },
            format="json",
        )
        self.assertEqual(invalid_uuid.status_code, 400)
        self.assertIn("invalid", invalid_uuid.data["error"])

        with patch("apps.scheduling.roster_views.MAX_ROSTER_ROWS", 1):
            too_many = self.client.patch(
                f"/events/roster/bulk?code={self.event.code}",
                {
                    "participantIds": [str(item.pk) for item in participants],
                    "updates": {"group": "X"},
                    "idempotencyKey": str(uuid.uuid4()),
                },
                format="json",
            )
        self.assertEqual(too_many.status_code, 400)

        first = participants[0]
        Weight.objects.filter(participant=first).delete()
        update_key = uuid.uuid4()
        update_payload = {
            "participantIds": [str(item.pk) for item in participants],
            "updates": {"groupName": "Unified", "weight": 0.4, "included": False},
            "idempotencyKey": str(update_key),
        }
        changed = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            update_payload,
            format="json",
        )
        self.assertEqual(changed.status_code, 200, changed.data)
        self.assertFalse(changed.data["idempotent"])
        self.assertEqual(changed.data["updatedCount"], 2)
        self.assertEqual(
            sorted(Weight.objects.filter(event=self.event).values_list("weight", flat=True)),
            [0.4, 0.4],
        )
        revision = changed.data["resultsRevision"]
        replay = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            update_payload,
            format="json",
        )
        self.assertEqual(replay.status_code, 200, replay.data)
        self.assertTrue(replay.data["idempotent"])
        self.assertEqual(replay.data["updatedCount"], 2)
        self.assertEqual(replay.data["resultsRevision"], revision)
        self.assertEqual(
            RosterBulkUpdateReceipt.objects.filter(
                event=self.event,
                idempotency_key=update_key,
            ).count(),
            1,
        )
        conflict = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                **update_payload,
                "updates": {"weight": 0.8},
            },
            format="json",
        )
        self.assertEqual(conflict.status_code, 409)

        ungrouped = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                "group": "Unified",
                "updates": {"group": ""},
                "idempotencyKey": str(uuid.uuid4()),
            },
            format="json",
        )
        self.assertEqual(ungrouped.status_code, 200)
        filtered = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                "filter": {"group": "__ungrouped__"},
                "updates": {"included": True},
                "idempotencyKey": str(uuid.uuid4()),
            },
            format="json",
        )
        self.assertEqual(filtered.status_code, 200)
        no_change = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                "group": "",
                "updates": {"group": "", "weight": 0.4, "included": True},
                "idempotencyKey": str(uuid.uuid4()),
            },
            format="json",
        )
        self.assertEqual(no_change.status_code, 200, no_change.data)
        self.assertEqual(no_change.data["updatedCount"], 0)

        for filter_data, message in [
            ({}, "explicit all=true"),
            ({"unknown": "value"}, "Unknown roster filter"),
            ({"all": False}, "filter.all must be true"),
        ]:
            with self.subTest(filter_data=filter_data):
                unsafe = self.client.patch(
                    f"/events/roster/bulk?code={self.event.code}",
                    {
                        "filter": filter_data,
                        "updates": {"group": "Unsafe"},
                        "idempotencyKey": str(uuid.uuid4()),
                    },
                    format="json",
                )
                self.assertEqual(unsafe.status_code, 400, unsafe.data)
                self.assertIn(message, unsafe.data["error"])
        all_selected = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                "filter": {"all": True},
                "updates": {"group": "Everyone"},
                "idempotencyKey": str(uuid.uuid4()),
            },
            format="json",
        )
        self.assertEqual(all_selected.status_code, 200, all_selected.data)
        self.assertEqual(all_selected.data["matchedCount"], 2)

        self.event.status = Event.Status.ARCHIVED
        self.event.save(update_fields=["status", "updated_at"])
        blocked = self.client.patch(
            f"/events/roster/bulk?code={self.event.code}",
            {
                "group": "",
                "updates": {"included": True},
                "idempotencyKey": str(uuid.uuid4()),
            },
            format="json",
        )
        self.assertEqual(blocked.status_code, 409)

    def test_model_strings_for_new_roster_entities(self):
        batch = self.preview()
        row = batch.rows.get(row_number=2)
        self.assertIn(self.event.code, str(batch))
        self.assertIn("Pasted data:2", str(row))
        receipt_response = self.commit(batch)
        receipt = RosterImportReceipt.objects.get(pk=receipt_response.data["receipt"]["id"])
        self.assertIn(self.event.code, str(receipt))
        invalidation = EventResultInvalidation.objects.create(event=self.event)
        self.assertIn("pending", str(invalidation))
        invalidation.processed_at = timezone.now()
        invalidation.save(update_fields=["processed_at", "updated_at"])
        self.assertIn("processed", str(invalidation))
        bulk_receipt = RosterBulkUpdateReceipt.objects.create(
            event=self.event,
            idempotency_key=uuid.uuid4(),
            request_fingerprint="a" * 64,
            matched_count=1,
            updated_count=1,
            results_revision=self.event.results_revision,
        )
        self.assertIn("roster bulk", str(bulk_receipt))
