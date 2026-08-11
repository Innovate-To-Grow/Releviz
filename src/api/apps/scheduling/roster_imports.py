from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import uuid
import zipfile
from collections import defaultdict
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.utils import timezone

from apps.authn.models import ContactEmail, EmailAuthChallenge
from apps.mail.models import EmailDeliveryJob, EmailMessageLog
from apps.scheduling.models import (
    Event,
    EventInvitation,
    EventResultSnapshot,
    Participant,
    RosterImportBatch,
    RosterImportReceipt,
    RosterImportRow,
    TemporaryEventSession,
    UserEvent,
    Weight,
)
from apps.scheduling.utils import default_availability

MAX_UPLOAD_BYTES = settings.ROSTER_IMPORT_MAX_FILE_BYTES
MAX_UNCOMPRESSED_BYTES = settings.ROSTER_IMPORT_MAX_UNCOMPRESSED_BYTES
MAX_COLUMNS = settings.ROSTER_IMPORT_MAX_COLUMNS
MAX_ROSTER_ROWS = settings.ROSTER_IMPORT_MAX_ROWS
MAX_PREVIEW_ROWS = MAX_ROSTER_ROWS * 5
PREVIEW_LIFETIME = settings.ROSTER_IMPORT_PREVIEW_LIFETIME

_HEADER_ALIASES = {
    "name": {"name", "full name", "participant", "participant name", "attendee"},
    "email": {"email", "email address", "e-mail", "e-mail address"},
    "group": {"group", "group name", "department", "cohort", "team"},
    "weight": {"weight", "priority"},
    "included": {"included", "include", "counted", "enabled"},
}
_MAPPING_KEYS = set(_HEADER_ALIASES)


class RosterImportError(ValueError):
    def __init__(self, message: str, *, status_code: int = 400, extra: dict | None = None):
        super().__init__(message)
        self.status_code = status_code
        self.extra = extra or {}


def _trim_row(values: list) -> list:
    values = list(values)
    while values and (values[-1] is None or values[-1] == ""):
        values.pop()
    return values


def _serialized_cell(cell):
    if getattr(cell, "data_type", "") == "f":
        return {"formula": str(cell.value or "")}
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, int | float | bool | str):
        return value
    return str(value)


def _parse_delimited(payload: bytes, *, worksheet: str) -> list[dict]:
    if len(payload) > MAX_UNCOMPRESSED_BYTES:
        raise RosterImportError("The table exceeds the 25 MiB uncompressed limit.")
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise RosterImportError("CSV and pasted data must be UTF-8 encoded.") from exc
    if "\x00" in text:
        raise RosterImportError("The table contains unsupported null bytes.")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample else ","
    try:
        rows = []
        for row_number, raw in enumerate(csv.reader(io.StringIO(text), delimiter=delimiter), 1):
            values = _trim_row([str(value) for value in raw])
            if not values or not any(str(value).strip() for value in values):
                continue
            if len(values) > MAX_COLUMNS:
                raise RosterImportError(f"Row {row_number} has more than {MAX_COLUMNS} columns.")
            rows.append(
                {
                    "worksheet": worksheet,
                    "row_number": row_number,
                    "raw_values": values,
                }
            )
            if len(rows) > MAX_PREVIEW_ROWS:
                raise RosterImportError(
                    f"A preview may contain at most {MAX_PREVIEW_ROWS} non-empty rows."
                )
    except csv.Error as exc:
        raise RosterImportError("The delimited table could not be parsed.") from exc
    if not rows:
        raise RosterImportError("The table is empty.")
    return rows


def _parse_xlsx(payload: bytes) -> list[dict]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            uncompressed_size = sum(entry.file_size for entry in archive.infolist())
            if uncompressed_size > MAX_UNCOMPRESSED_BYTES:
                raise RosterImportError("The workbook exceeds the 25 MiB uncompressed limit.")
            if any(entry.flag_bits & 0x1 for entry in archive.infolist()):
                raise RosterImportError("Encrypted workbooks are not supported.")
    except zipfile.BadZipFile as exc:
        raise RosterImportError("The uploaded file is not a valid .xlsx workbook.") from exc

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is required in deployed builds
        raise RosterImportError("Excel import support is unavailable.", status_code=503) from exc

    try:
        workbook = load_workbook(
            io.BytesIO(payload),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise RosterImportError("The uploaded file is not a readable .xlsx workbook.") from exc

    parsed = []
    try:
        for worksheet in workbook.worksheets:
            if worksheet.max_column > MAX_COLUMNS:
                raise RosterImportError(
                    f"Worksheet '{worksheet.title}' declares more than {MAX_COLUMNS} columns."
                )
            if worksheet.max_row > MAX_PREVIEW_ROWS + 1:
                raise RosterImportError(
                    f"Worksheet '{worksheet.title}' declares more than {MAX_PREVIEW_ROWS + 1} rows."
                )
            for row_number, cells in enumerate(
                worksheet.iter_rows(
                    max_row=MAX_PREVIEW_ROWS + 1,
                    max_col=MAX_COLUMNS,
                ),
                1,
            ):
                values = _trim_row([_serialized_cell(cell) for cell in cells])
                if not values or not any(value != "" for value in values):
                    continue
                if len(values) > MAX_COLUMNS:
                    raise RosterImportError(
                        f"Worksheet '{worksheet.title}' row {row_number} has more than "
                        f"{MAX_COLUMNS} columns."
                    )
                parsed.append(
                    {
                        "worksheet": worksheet.title,
                        "row_number": row_number,
                        "raw_values": values,
                    }
                )
                if len(parsed) > MAX_PREVIEW_ROWS:
                    raise RosterImportError(
                        f"A workbook preview may contain at most {MAX_PREVIEW_ROWS} non-empty rows."
                    )
    finally:
        workbook.close()
    if not parsed:
        raise RosterImportError("The workbook contains no non-empty rows.")
    return parsed


def parse_roster_source(*, uploaded_file=None, pasted_text=None, requested_source=""):
    if uploaded_file is not None:
        if getattr(uploaded_file, "size", 0) > MAX_UPLOAD_BYTES:
            raise RosterImportError("The uploaded file exceeds the 5 MiB limit.")
        payload = uploaded_file.read(MAX_UPLOAD_BYTES + 1)
        if len(payload) > MAX_UPLOAD_BYTES:
            raise RosterImportError("The uploaded file exceeds the 5 MiB limit.")
        extension = Path(str(uploaded_file.name or "")).suffix.lower()
        if extension == ".xlsx":
            return RosterImportBatch.SourceType.XLSX, "upload.xlsx", _parse_xlsx(payload)
        if extension == ".csv":
            return (
                RosterImportBatch.SourceType.CSV,
                "upload.csv",
                _parse_delimited(payload, worksheet="CSV"),
            )
        if extension == ".xls":
            raise RosterImportError("Legacy .xls files are not supported; save the file as .xlsx.")
        raise RosterImportError("Upload a .csv or .xlsx file.")

    if requested_source and requested_source != RosterImportBatch.SourceType.PASTE:
        raise RosterImportError("sourceType must be 'paste' when no file is uploaded.")
    if not isinstance(pasted_text, str) or not pasted_text.strip():
        raise RosterImportError("pastedText is required when no file is uploaded.")
    payload = pasted_text.encode("utf-8")
    return (
        RosterImportBatch.SourceType.PASTE,
        "pasted table",
        _parse_delimited(payload, worksheet="Pasted data"),
    )


def _display_cell(value) -> str:
    if isinstance(value, dict) and "formula" in value:
        return f"={value['formula']}"
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _canonical_header(value) -> str:
    return " ".join(_display_cell(value).lower().replace("_", " ").split())


def _auto_mapping(headers: list) -> dict:
    mapping = {}
    for index, header in enumerate(headers):
        canonical = _canonical_header(header)
        for field, aliases in _HEADER_ALIASES.items():
            if field not in mapping and canonical in aliases:
                mapping[field] = index
    return mapping


def _worksheet_metadata(raw_rows: list[dict]) -> list[dict]:
    grouped = defaultdict(list)
    for row in raw_rows:
        grouped[row["worksheet"]].append(row)
    metadata = []
    for worksheet, rows in grouped.items():
        rows.sort(key=lambda item: item["row_number"])
        first = rows[0]
        metadata.append(
            {
                "name": worksheet,
                "rowCount": max(len(rows) - 1, 0),
                "columnCount": max(len(row["raw_values"]) for row in rows),
                "headers": [_display_cell(value) for value in first["raw_values"]],
                "defaultHeaderRow": first["row_number"],
            }
        )
    return metadata


def _mapping_index(value, headers: list, field: str) -> int:
    if isinstance(value, bool):
        raise RosterImportError(f"columnMapping.{field} must identify a column.")
    if isinstance(value, int):
        index = value
    elif isinstance(value, str) and value.strip().isdigit():
        index = int(value.strip())
    elif isinstance(value, str):
        matches = [
            index
            for index, header in enumerate(headers)
            if _canonical_header(header) == _canonical_header(value)
        ]
        if len(matches) != 1:
            raise RosterImportError(
                f"columnMapping.{field} must match exactly one header or use a zero-based index."
            )
        index = matches[0]
    else:
        raise RosterImportError(f"columnMapping.{field} must identify a column.")
    if index < 0 or index >= len(headers) or index >= MAX_COLUMNS:
        raise RosterImportError(f"columnMapping.{field} is outside the available columns.")
    return index


def _parse_defaults(value) -> dict:
    if value is None:
        return {"group": "", "weight": 1.0, "included": True}
    if not isinstance(value, dict):
        raise RosterImportError("defaults must be an object.")
    group_name = str(value.get("group", value.get("groupName", "")) or "").strip()
    if len(group_name) > 100:
        raise RosterImportError("defaults.group is too long (max 100).")
    try:
        weight = float(value.get("weight", 1.0))
    except (TypeError, ValueError) as exc:
        raise RosterImportError("defaults.weight must be between 0 and 1.") from exc
    if not math.isfinite(weight) or weight < 0 or weight > 1:
        raise RosterImportError("defaults.weight must be between 0 and 1.")
    included = _parse_included(value.get("included", True), label="defaults.included")
    return {"group": group_name, "weight": weight, "included": included}


def _parse_included(value, *, label="included") -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    normalized = str(value or "").strip().lower()
    if normalized in {"1", "true", "yes", "y", "included", "include"}:
        return True
    if normalized in {"0", "false", "no", "n", "excluded", "exclude"}:
        return False
    raise RosterImportError(f"{label} must be true or false.")


def _mapped_value(row: RosterImportRow, mapping: dict, field: str):
    if field not in mapping:
        return None, None
    index = mapping[field]
    value = row.raw_values[index] if index < len(row.raw_values) else ""
    if isinstance(value, dict) and "formula" in value:
        return None, f"{field} cannot contain a formula."
    if isinstance(value, str) and value.lstrip().startswith("="):
        return None, f"{field} cannot contain a formula."
    return value, None


def _validate_identity_fields(name: str, email: str, group_name: str) -> list[str]:
    errors = []
    if not name:
        errors.append("name is required.")
    elif len(name) > 100:
        errors.append("name is too long (max 100).")
    if not email:
        errors.append("email is required.")
    elif len(email) > 254:
        errors.append("email is too long (max 254).")
    else:
        try:
            validate_email(email)
        except ValidationError:
            errors.append("email is invalid.")
    if len(group_name) > 100:
        errors.append("group is too long (max 100).")
    return errors


def _normalize_row(row: RosterImportRow, mapping: dict, defaults: dict) -> None:
    errors = []
    raw_name, error = _mapped_value(row, mapping, "name")
    if error:
        errors.append(error)
    raw_email, error = _mapped_value(row, mapping, "email")
    if error:
        errors.append(error)
    if "name" not in mapping:
        errors.append("Map a name column.")
    if "email" not in mapping:
        errors.append("Map an email column.")

    raw_group, error = _mapped_value(row, mapping, "group")
    if error:
        errors.append(error)
    raw_weight, error = _mapped_value(row, mapping, "weight")
    if error:
        errors.append(error)
    raw_included, error = _mapped_value(row, mapping, "included")
    if error:
        errors.append(error)

    name = _display_cell(raw_name)
    email = _display_cell(raw_email).lower()
    group_name = _display_cell(raw_group) if "group" in mapping else str(defaults.get("group", ""))
    weight = defaults.get("weight", 1.0)
    if "weight" in mapping and raw_weight not in {None, ""}:
        try:
            weight = float(raw_weight)
            if not math.isfinite(weight) or weight < 0 or weight > 1:
                raise ValueError
        except (TypeError, ValueError):
            errors.append("weight must be between 0 and 1.")
            weight = 1.0
    included = defaults.get("included", True)
    if "included" in mapping and raw_included not in {None, ""}:
        try:
            included = _parse_included(raw_included)
        except RosterImportError as exc:
            errors.append(str(exc))
            included = True

    errors.extend(_validate_identity_fields(name, email, group_name))
    row.name = name[:100]
    row.email = email[:254]
    row.group_name = group_name[:100]
    row.weight = weight
    row.included = included
    row.selected = True
    row.validation_errors = list(dict.fromkeys(errors))
    row.duplicate_status = RosterImportRow.DuplicateStatus.UNIQUE


def _remove_duplicate_error(errors: list) -> list:
    return [error for error in errors if error != "Conflicting duplicate email."]


def _apply_duplicate_rules(rows: list[RosterImportRow]) -> None:
    by_email = defaultdict(list)
    for row in rows:
        row.validation_errors = _remove_duplicate_error(row.validation_errors or [])
        if row.selected:
            row.duplicate_status = RosterImportRow.DuplicateStatus.UNIQUE
            if row.email:
                by_email[row.email].append(row)
        elif row.duplicate_status == RosterImportRow.DuplicateStatus.CONFLICT:
            row.duplicate_status = RosterImportRow.DuplicateStatus.UNIQUE

    for duplicates in by_email.values():
        if len(duplicates) < 2:
            continue
        signatures = {
            (row.name, row.email, row.group_name, float(row.weight), bool(row.included))
            for row in duplicates
        }
        if len(signatures) == 1:
            for duplicate in duplicates[1:]:
                duplicate.selected = False
                duplicate.duplicate_status = RosterImportRow.DuplicateStatus.IDENTICAL
            continue
        for duplicate in duplicates:
            duplicate.duplicate_status = RosterImportRow.DuplicateStatus.CONFLICT
            duplicate.validation_errors = list(
                dict.fromkeys(
                    [*(duplicate.validation_errors or []), "Conflicting duplicate email."]
                )
            )


def _summary(rows: list[RosterImportRow]) -> dict:
    selected = [row for row in rows if row.selected]
    valid = [row for row in selected if not row.validation_errors]
    conflicts = [
        row for row in selected if row.duplicate_status == RosterImportRow.DuplicateStatus.CONFLICT
    ]
    return {
        "total": len(rows),
        "selected": len(selected),
        "valid": len(valid),
        "invalid": len(selected) - len(valid),
        "conflicts": len(conflicts),
    }


def _active_rows(batch: RosterImportBatch) -> list[RosterImportRow]:
    if not batch.selected_worksheet:
        return []
    return list(
        batch.rows.filter(
            worksheet=batch.selected_worksheet,
            row_number__gt=batch.header_row,
        ).order_by("row_number")
    )


def normalize_import_batch(batch: RosterImportBatch) -> None:
    if not batch.selected_worksheet:
        batch.summary = {
            "total": 0,
            "selected": 0,
            "valid": 0,
            "invalid": 0,
            "conflicts": 0,
        }
        batch.save(update_fields=["summary", "updated_at"])
        return
    batch.rows.exclude(worksheet=batch.selected_worksheet).update(selected=False)
    rows = _active_rows(batch)
    for row in rows:
        _normalize_row(row, batch.column_mapping or {}, batch.defaults or {})
    _apply_duplicate_rules(rows)
    if _summary(rows)["valid"] > MAX_ROSTER_ROWS:
        raise RosterImportError(
            f"An import may contain at most {MAX_ROSTER_ROWS} valid participants."
        )
    if rows:
        RosterImportRow.objects.bulk_update(
            rows,
            [
                "name",
                "email",
                "group_name",
                "weight",
                "included",
                "selected",
                "validation_errors",
                "duplicate_status",
                "updated_at",
            ],
        )
    batch.summary = _summary(rows)
    batch.save(update_fields=["summary", "updated_at"])


def _header_values(batch: RosterImportBatch, worksheet: str, header_row: int) -> list:
    row = batch.rows.filter(worksheet=worksheet, row_number=header_row).first()
    if row is None:
        raise RosterImportError("headerRow does not identify a non-empty row in the worksheet.")
    return row.raw_values


def create_roster_import(
    *,
    event: Event,
    created_by,
    uploaded_file=None,
    pasted_text=None,
    requested_source="",
) -> RosterImportBatch:
    source_type, source_label, raw_rows = parse_roster_source(
        uploaded_file=uploaded_file,
        pasted_text=pasted_text,
        requested_source=requested_source,
    )
    worksheet_metadata = _worksheet_metadata(raw_rows)
    selected = worksheet_metadata[0]["name"] if len(worksheet_metadata) == 1 else ""
    header_row = worksheet_metadata[0]["defaultHeaderRow"] if selected else 1
    headers = worksheet_metadata[0]["headers"] if selected else []
    mapping = _auto_mapping(headers) if selected else {}
    with transaction.atomic():
        batch = RosterImportBatch.objects.create(
            event=event,
            created_by=created_by,
            source_type=source_type,
            source_label=source_label,
            worksheets=worksheet_metadata,
            selected_worksheet=selected,
            header_row=header_row,
            column_mapping=mapping,
            defaults={"group": "", "weight": 1.0, "included": True},
            expires_at=timezone.now() + PREVIEW_LIFETIME,
        )
        RosterImportRow.objects.bulk_create(
            [
                RosterImportRow(
                    batch=batch,
                    worksheet=row["worksheet"],
                    row_number=row["row_number"],
                    raw_values=row["raw_values"],
                )
                for row in raw_rows
            ]
        )
        normalize_import_batch(batch)
    return batch


def update_roster_import(*, batch: RosterImportBatch, data) -> RosterImportBatch:
    with transaction.atomic():
        batch = RosterImportBatch.objects.select_for_update().get(pk=batch.pk)
        _require_preview(batch)
        mapping_changed = False
        if "worksheet" in data:
            worksheet = str(data.get("worksheet") or "")
            if worksheet not in {item["name"] for item in batch.worksheets}:
                raise RosterImportError("worksheet was not found in this import.")
            batch.selected_worksheet = worksheet
            first_row = batch.rows.filter(worksheet=worksheet).order_by("row_number").first()
            if first_row is None:
                raise RosterImportError("worksheet contains no non-empty rows.")
            if "headerRow" not in data:
                batch.header_row = first_row.row_number
            mapping_changed = True
        if "headerRow" in data:
            header_row = data.get("headerRow")
            if isinstance(header_row, bool) or not isinstance(header_row, int) or header_row < 1:
                raise RosterImportError("headerRow must be a positive integer.")
            batch.header_row = header_row
            mapping_changed = True
        if not batch.selected_worksheet:
            raise RosterImportError("Select a worksheet before mapping columns.")
        headers = _header_values(batch, batch.selected_worksheet, batch.header_row)

        if "columnMapping" in data:
            supplied = data.get("columnMapping")
            if not isinstance(supplied, dict):
                raise RosterImportError("columnMapping must be an object.")
            unknown = set(supplied) - _MAPPING_KEYS
            if unknown:
                raise RosterImportError(f"Unknown mapped field: {sorted(unknown)[0]}.")
            mapping = {
                field: _mapping_index(value, headers, field)
                for field, value in supplied.items()
                if value is not None and value != ""
            }
            if "name" not in mapping or "email" not in mapping:
                raise RosterImportError("columnMapping must include name and email.")
            if len(set(mapping.values())) != len(mapping):
                raise RosterImportError("Each mapped field must use a different column.")
            batch.column_mapping = mapping
            mapping_changed = True
        elif mapping_changed:
            batch.column_mapping = _auto_mapping(headers)

        if "defaults" in data:
            batch.defaults = _parse_defaults(data.get("defaults"))
            mapping_changed = True
        batch.save(
            update_fields=[
                "selected_worksheet",
                "header_row",
                "column_mapping",
                "defaults",
                "updated_at",
            ]
        )
        if mapping_changed:
            normalize_import_batch(batch)
        if "rowUpdates" in data:
            _apply_row_updates(batch, data.get("rowUpdates"))
    return batch


def _apply_row_updates(batch: RosterImportBatch, updates) -> None:
    if not isinstance(updates, list):
        raise RosterImportError("rowUpdates must be an array.")
    if len(updates) > MAX_ROSTER_ROWS:
        raise RosterImportError(f"rowUpdates may contain at most {MAX_ROSTER_ROWS} rows.")
    identifiers = [str(item.get("id") or "") for item in updates if isinstance(item, dict)]
    if len(identifiers) != len(updates) or len(set(identifiers)) != len(identifiers):
        raise RosterImportError("Each rowUpdates entry must have a unique id.")
    all_rows = _active_rows(batch)
    rows = {str(row.pk): row for row in all_rows if str(row.pk) in set(identifiers)}
    if len(rows) != len(updates):
        raise RosterImportError("A rowUpdates id does not belong to this preview.")

    for item in updates:
        row = rows[str(item["id"])]
        name = row.name
        email = row.email
        group_name = row.group_name
        if "name" in item:
            name = str(item.get("name") or "").strip()
        if "email" in item:
            email = str(item.get("email") or "").strip().lower()
        if "group" in item or "groupName" in item:
            group_name = str(item.get("group", item.get("groupName")) or "").strip()
        identity_errors = _validate_identity_fields(name, email, group_name)
        row.name = name[:100]
        row.email = email[:254]
        row.group_name = group_name[:100]
        if "weight" in item:
            try:
                weight = float(item.get("weight"))
            except (TypeError, ValueError) as exc:
                raise RosterImportError("rowUpdates.weight must be between 0 and 1.") from exc
            if not math.isfinite(weight) or weight < 0 or weight > 1:
                raise RosterImportError("rowUpdates.weight must be between 0 and 1.")
            row.weight = weight
        if "included" in item:
            row.included = _parse_included(item.get("included"), label="rowUpdates.included")
        if "selected" in item:
            row.selected = _parse_included(item.get("selected"), label="rowUpdates.selected")
        row.validation_errors = identity_errors
        row.duplicate_status = RosterImportRow.DuplicateStatus.UNIQUE

    _apply_duplicate_rules(all_rows)
    if _summary(all_rows)["valid"] > MAX_ROSTER_ROWS:
        raise RosterImportError(
            f"An import may contain at most {MAX_ROSTER_ROWS} valid participants."
        )
    RosterImportRow.objects.bulk_update(
        all_rows,
        [
            "name",
            "email",
            "group_name",
            "weight",
            "included",
            "selected",
            "validation_errors",
            "duplicate_status",
            "updated_at",
        ],
    )
    batch.summary = _summary(all_rows)
    batch.save(update_fields=["summary", "updated_at"])


def _require_preview(batch: RosterImportBatch) -> None:
    if batch.status != RosterImportBatch.Status.PREVIEW:
        raise RosterImportError(
            f"This import is {batch.status} and can no longer be changed.",
            status_code=409,
        )
    if batch.expires_at <= timezone.now():
        _scrub_batch(batch, RosterImportBatch.Status.EXPIRED)
        raise RosterImportError("This import preview has expired.", status_code=410)


def _scrub_batch(batch: RosterImportBatch, status: str, *, summary=None) -> None:
    batch.rows.all().delete()
    batch.status = status
    batch.worksheets = []
    batch.selected_worksheet = ""
    batch.column_mapping = {}
    batch.defaults = {}
    batch.summary = summary or {}
    batch.failure_reason = ""
    batch.save(
        update_fields=[
            "status",
            "worksheets",
            "selected_worksheet",
            "column_mapping",
            "defaults",
            "summary",
            "failure_reason",
            "updated_at",
        ]
    )


def expire_roster_import_preview(batch: RosterImportBatch) -> bool:
    """Expire one requested preview without sweeping unrelated organizers' data."""

    with transaction.atomic():
        batch = RosterImportBatch.objects.select_for_update().get(pk=batch.pk)
        if batch.status == RosterImportBatch.Status.EXPIRED:
            return True
        if batch.status == RosterImportBatch.Status.PREVIEW and batch.expires_at <= timezone.now():
            _scrub_batch(batch, RosterImportBatch.Status.EXPIRED)
            return True
    return False


def expire_stale_roster_imports(*, limit: int | None = None) -> int:
    stale_query = (
        RosterImportBatch.objects.filter(
            status=RosterImportBatch.Status.PREVIEW,
            expires_at__lte=timezone.now(),
        )
        .order_by("expires_at")
        .values_list("pk", flat=True)
    )
    stale = list(stale_query[:limit] if limit is not None else stale_query)
    for batch_id in stale:
        with transaction.atomic():
            batch = RosterImportBatch.objects.select_for_update().filter(pk=batch_id).first()
            if (
                batch is not None
                and batch.status == RosterImportBatch.Status.PREVIEW
                and batch.expires_at <= timezone.now()
            ):
                _scrub_batch(batch, RosterImportBatch.Status.EXPIRED)
    return len(stale)


def cancel_roster_import(batch: RosterImportBatch) -> RosterImportBatch:
    with transaction.atomic():
        batch = RosterImportBatch.objects.select_for_update().get(pk=batch.pk)
        if batch.status == RosterImportBatch.Status.COMMITTED:
            raise RosterImportError("A committed import cannot be canceled.", status_code=409)
        if batch.status == RosterImportBatch.Status.CANCELED:
            return batch
        _scrub_batch(batch, RosterImportBatch.Status.CANCELED)
    return batch


def _fingerprint(payload: dict) -> str:
    serialized = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def _resolve_members(rows: list[RosterImportRow]):
    emails = [row.email for row in rows]
    Member = get_user_model()
    initial_contacts = list(
        ContactEmail.objects.filter(email_address__in=emails).only(
            "pk",
            "email_address",
            "member_id",
        )
    )
    initial_member_ids = sorted(
        {contact.member_id for contact in initial_contacts if contact.member_id is not None},
        key=str,
    )
    locked_members = {
        member.pk: member
        for member in Member.objects.select_for_update()
        .filter(pk__in=initial_member_ids)
        .order_by("pk")
    }
    locked_contacts = list(
        ContactEmail.objects.select_for_update()
        .filter(email_address__in=emails)
        .order_by("email_address")
    )
    if any(
        contact.member_id is not None and contact.member_id not in locked_members
        for contact in locked_contacts
    ):
        raise RosterImportError(
            "An account changed while the roster was being imported; retry the commit.",
            status_code=409,
        )
    contacts = {}
    for contact in locked_contacts:
        if contact.member_id is not None:
            contact.member = locked_members[contact.member_id]
        contacts[contact.email_address.lower()] = contact
    new_members = []
    new_contacts = []
    orphan_contacts = []
    resolved = {}
    names = {row.email: row.name for row in rows}

    for email in emails:
        contact = contacts.get(email)
        if contact is not None and contact.member_id is not None:
            member = contact.member
            if not member.is_active:
                raise RosterImportError(
                    "A roster email belongs to an inactive account.",
                    status_code=409,
                )
            if getattr(member, "access_level", "full") == "full" and not contact.verified:
                raise RosterImportError(
                    "A roster email belongs to an unverified full account.",
                    status_code=409,
                )
            resolved[email] = member
            continue

        member = Member(
            email=email,
            first_name=names[email],
            is_active=True,
            access_level=Member.AccessLevel.TEMPORARY,
        )
        member.set_unusable_password()
        new_members.append(member)
        resolved[email] = member
        if contact is None:
            new_contacts.append(
                ContactEmail(
                    member=member,
                    email_address=email,
                    email_type="primary",
                    verified=False,
                )
            )
        else:
            contact.member = member
            contact.email_type = "primary"
            contact.verified = False
            orphan_contacts.append(contact)

    if new_members:
        Member.objects.bulk_create(new_members)
    if new_contacts:
        ContactEmail.objects.bulk_create(new_contacts)
    if orphan_contacts:
        ContactEmail.objects.bulk_update(
            orphan_contacts,
            ["member", "email_type", "verified", "updated_at"],
        )
    member_ids = [member.pk for member in resolved.values()]
    if len(set(member_ids)) != len(member_ids):
        raise RosterImportError(
            "Two roster email addresses resolve to the same account.",
            status_code=409,
        )
    return resolved


def _rebuild_event_roster(event: Event, now) -> None:
    # Keep the destructive path in the same lock order as account upgrades:
    # member/contact (resolved by the caller), participant, challenge/job,
    # session, then invitation. Schedule writes that already own one
    # participant can finish without forming Participant <-> Invitation cycles.
    participants = list(event.participants.select_for_update().order_by("pk"))
    challenge_scope_prefix = f"temp-event:{event.pk}:invitation:"
    challenges = list(
        EmailAuthChallenge.objects.select_for_update()
        .filter(
            purpose=EmailAuthChallenge.Purpose.TEMP_EVENT_ACCESS,
            scope_key__startswith=challenge_scope_prefix,
        )
        .order_by("pk")
    )
    challenge_ids = [challenge.pk for challenge in challenges]
    jobs = list(
        EmailDeliveryJob.objects.select_for_update()
        .filter(Q(event=event) | Q(auth_challenge_id__in=challenge_ids))
        .order_by("pk")
    )
    if any(job.status == EmailDeliveryJob.Status.PROCESSING for job in jobs):
        raise RosterImportError(
            "Wait for in-progress email deliveries to finish before rebuilding the roster.",
            status_code=409,
        )
    if challenge_ids:
        EmailAuthChallenge.objects.filter(
            pk__in=challenge_ids,
            status=EmailAuthChallenge.Status.PENDING,
        ).update(status=EmailAuthChallenge.Status.EXPIRED, updated_at=now)
        EmailDeliveryJob.objects.filter(
            auth_challenge_id__in=challenge_ids,
            status__in=[
                EmailDeliveryJob.Status.PENDING,
                EmailDeliveryJob.Status.RETRY,
            ],
        ).update(
            status=EmailDeliveryJob.Status.CANCELED,
            last_error="The event roster was rebuilt.",
            locked_at=None,
            lock_token=None,
            updated_at=now,
        )
    sessions = list(
        TemporaryEventSession.objects.select_for_update()
        .filter(participant__event=event, revoked_at__isnull=True)
        .order_by("pk")
    )
    if sessions:
        TemporaryEventSession.objects.filter(pk__in=[session.pk for session in sessions]).update(
            revoked_at=now,
            updated_at=now,
        )
    EmailDeliveryJob.objects.filter(
        event=event,
        status__in=[
            EmailDeliveryJob.Status.PENDING,
            EmailDeliveryJob.Status.RETRY,
        ],
    ).exclude(message_type=EmailMessageLog.MessageType.FINAL_CANCELLATION).update(
        status=EmailDeliveryJob.Status.CANCELED,
        locked_at=None,
        lock_token=None,
        updated_at=now,
    )
    event.user_events.filter(role="participant").delete()
    event.roster_bulk_update_receipts.all().delete()
    event.email_delivery_requests.exclude(
        operation="final_cancellation",
    ).delete()
    invitations = list(event.invitations.select_for_update().order_by("pk"))
    if invitations:
        EventInvitation.objects.filter(
            pk__in=[invitation.pk for invitation in invitations]
        ).delete()
    if participants:
        Participant.objects.filter(pk__in=[participant.pk for participant in participants]).delete()
    event.status = Event.Status.DRAFT
    event.opened_at = None
    event.finalized_at = None
    event.closed_at = None
    event.archived_at = None
    event.version += 1


def _write_roster(
    event: Event,
    organizer,
    rows: list[RosterImportRow],
    *,
    members=None,
):
    members = members or _resolve_members(rows)
    member_ids = [members[row.email].pk for row in rows]
    existing = {
        participant.member_id: participant
        for participant in Participant.objects.select_for_update().filter(
            event=event,
            member_id__in=member_ids,
        )
    }
    current_member_ids = set(
        Participant.objects.filter(event=event).values_list("member_id", flat=True)
    )
    if len(current_member_ids | set(member_ids)) > MAX_ROSTER_ROWS:
        raise RosterImportError(
            f"An event can have at most {MAX_ROSTER_ROWS} participants.",
            status_code=409,
        )
    roster_emails = {row.email for row in rows}
    current_invitation_emails = set(event.invitations.values_list("email", flat=True))
    if len(current_invitation_emails | roster_emails) > MAX_ROSTER_ROWS:
        raise RosterImportError(
            f"An event can have at most {MAX_ROSTER_ROWS} invitation recipients.",
            status_code=409,
        )

    new_participants = []
    changed_participants = []
    participant_by_email = {}
    for sort_order, row in enumerate(rows, 1):
        member = members[row.email]
        participant = existing.get(member.pk)
        if participant is None:
            participant = Participant(
                event=event,
                member=member,
                participant_name=row.name,
                availability_inperson=default_availability(event),
                availability_virtual=default_availability(event),
                group_name=row.group_name or None,
                sort_order=sort_order,
            )
            new_participants.append(participant)
        else:
            changed = False
            values = {
                "participant_name": row.name,
                "group_name": row.group_name or None,
                "hidden": False,
            }
            for field, value in values.items():
                if getattr(participant, field) != value:
                    setattr(participant, field, value)
                    changed = True
            if changed:
                participant.version += 1
                changed_participants.append(participant)
        participant_by_email[row.email] = participant

    if new_participants:
        Participant.objects.bulk_create(new_participants)
    if changed_participants:
        Participant.objects.bulk_update(
            changed_participants,
            ["participant_name", "group_name", "hidden", "version", "updated_at"],
        )

    UserEvent.objects.bulk_create(
        [UserEvent(member=members[row.email], event=event, role="participant") for row in rows],
        ignore_conflicts=True,
    )

    invitations = {
        invitation.email: invitation
        for invitation in EventInvitation.objects.select_for_update().filter(
            event=event,
            email__in=[row.email for row in rows],
        )
    }
    new_invitations = []
    changed_invitations = []
    for row in rows:
        member = members[row.email]
        invitation = invitations.get(row.email)
        if invitation is None:
            new_invitations.append(
                EventInvitation(
                    event=event,
                    email=row.email,
                    member=member,
                    invited_by=organizer,
                )
            )
            continue
        changed = False
        if invitation.member_id != member.pk:
            invitation.member = member
            changed = True
        if invitation.invited_by_id is None:
            invitation.invited_by = organizer
            changed = True
        if changed:
            changed_invitations.append(invitation)
    if new_invitations:
        EventInvitation.objects.bulk_create(new_invitations)
    if changed_invitations:
        EventInvitation.objects.bulk_update(
            changed_invitations,
            ["member", "invited_by", "updated_at"],
        )

    participant_ids = [participant_by_email[row.email].pk for row in rows]
    weights = {
        weight.participant_id: weight
        for weight in Weight.objects.select_for_update().filter(
            event=event,
            participant_id__in=participant_ids,
        )
    }
    new_weights = []
    changed_weights = []
    for row in rows:
        participant = participant_by_email[row.email]
        weight = weights.get(participant.pk)
        if weight is None:
            new_weights.append(
                Weight(
                    event=event,
                    participant=participant,
                    weight=row.weight,
                    included=row.included,
                )
            )
        elif float(weight.weight) != float(row.weight) or weight.included != row.included:
            weight.weight = row.weight
            weight.included = row.included
            changed_weights.append(weight)
    if new_weights:
        Weight.objects.bulk_create(new_weights)
    if changed_weights:
        Weight.objects.bulk_update(changed_weights, ["weight", "included", "updated_at"])
    return len(new_participants), len(rows) - len(new_participants)


def commit_roster_import(*, event: Event, batch_id, organizer, data):
    mode = str(data.get("mode") or "").strip().lower()
    if mode not in RosterImportReceipt.Mode.values:
        raise RosterImportError("mode must be 'merge' or 'rebuild'.")
    try:
        idempotency_key = uuid.UUID(str(data.get("idempotencyKey") or ""))
    except (TypeError, ValueError, AttributeError) as exc:
        raise RosterImportError("idempotencyKey must be a UUID.") from exc
    fingerprint = _fingerprint({"batchId": str(batch_id), "mode": mode})

    try:
        with transaction.atomic():
            event = Event.objects.select_for_update().get(pk=event.pk)
            batch = (
                RosterImportBatch.objects.select_for_update()
                .filter(pk=batch_id, event=event)
                .first()
            )
            if batch is None:
                raise RosterImportError("Roster import not found.", status_code=404)
            previous = RosterImportReceipt.objects.filter(
                event=event,
                idempotency_key=idempotency_key,
            ).first()
            if previous is not None:
                if previous.request_fingerprint != fingerprint:
                    raise RosterImportError(
                        "This idempotency key was already used for a different import.",
                        status_code=409,
                    )
                return previous, True

            _require_preview(batch)
            if event.organizer_id != organizer.pk:
                raise RosterImportError(
                    "Only the organizer can commit a roster import.",
                    status_code=403,
                )
            if event.status in {Event.Status.FINALIZED, Event.Status.ARCHIVED}:
                raise RosterImportError(
                    "Reopen this event before changing its roster.",
                    status_code=409,
                )
            if getattr(event, "final_meeting", None) is not None and event.final_meeting.active:
                raise RosterImportError(
                    "Reopen this event before changing a roster with a confirmed meeting.",
                    status_code=409,
                )
            if mode == RosterImportReceipt.Mode.REBUILD:
                confirmation_code = str(data.get("confirmationCode") or "").strip().upper()
                if confirmation_code != event.code.upper():
                    raise RosterImportError(
                        "confirmationCode must match the event code for rebuild."
                    )

            rows = list(
                batch.rows.select_for_update()
                .filter(
                    worksheet=batch.selected_worksheet,
                    row_number__gt=batch.header_row,
                    selected=True,
                )
                .order_by("row_number")
            )
            if not rows:
                raise RosterImportError("Select at least one valid roster row.")
            invalid = [row for row in rows if row.validation_errors]
            if invalid:
                raise RosterImportError(
                    "Resolve or deselect invalid roster rows before committing.",
                    status_code=409,
                    extra={"invalidRowCount": len(invalid)},
                )
            if len(rows) > MAX_ROSTER_ROWS:
                raise RosterImportError(
                    f"An import may contain at most {MAX_ROSTER_ROWS} participants."
                )

            batch.status = RosterImportBatch.Status.COMMITTING
            batch.save(update_fields=["status", "updated_at"])
            now = timezone.now()
            members = _resolve_members(rows)
            if mode == RosterImportReceipt.Mode.REBUILD:
                _rebuild_event_roster(event, now)
            created_count, updated_count = _write_roster(
                event,
                organizer,
                rows,
                members=members,
            )
            event.results_revision += 1
            event_update_fields = ["results_revision", "updated_at"]
            if mode == RosterImportReceipt.Mode.REBUILD:
                event_update_fields.extend(
                    [
                        "status",
                        "opened_at",
                        "finalized_at",
                        "closed_at",
                        "archived_at",
                        "version",
                    ]
                )
            event.save(update_fields=event_update_fields)
            EventResultSnapshot.objects.update_or_create(
                event=event,
                defaults={
                    "requested_revision": event.results_revision,
                    "status": EventResultSnapshot.Status.REFRESHING,
                    "last_error": "",
                },
            )
            receipt = RosterImportReceipt.objects.create(
                event=event,
                batch=batch,
                committed_by=organizer,
                idempotency_key=idempotency_key,
                request_fingerprint=fingerprint,
                mode=mode,
                imported_count=len(rows),
                created_count=created_count,
                updated_count=updated_count,
                results_revision=event.results_revision,
                committed_at=now,
            )
            _scrub_batch(
                batch,
                RosterImportBatch.Status.COMMITTED,
                summary={
                    "imported": len(rows),
                    "created": created_count,
                    "updated": updated_count,
                },
            )
            return receipt, False
    except IntegrityError as exc:
        raise RosterImportError(
            "The roster changed concurrently; refresh the preview and try again.",
            status_code=409,
        ) from exc


def roster_import_payload(batch: RosterImportBatch) -> dict:
    selected_metadata = next(
        (item for item in batch.worksheets if item.get("name") == batch.selected_worksheet),
        None,
    )
    headers = []
    if batch.selected_worksheet and batch.status == RosterImportBatch.Status.PREVIEW:
        header = batch.rows.filter(
            worksheet=batch.selected_worksheet,
            row_number=batch.header_row,
        ).first()
        if header is not None:
            headers = [_display_cell(value) for value in header.raw_values]
    return {
        "id": str(batch.pk),
        "status": batch.status,
        "sourceType": batch.source_type,
        "fileName": batch.source_label or None,
        "worksheets": batch.worksheets,
        "selectedWorksheet": batch.selected_worksheet or None,
        "headerRow": batch.header_row,
        "headers": headers or (selected_metadata or {}).get("headers", []),
        "columnMapping": batch.column_mapping,
        "defaults": batch.defaults,
        "expiresAt": batch.expires_at.isoformat(),
        "summary": batch.summary,
    }


def roster_import_row_payload(row: RosterImportRow) -> dict:
    return {
        "id": str(row.pk),
        "rowNumber": row.row_number,
        "name": row.name,
        "email": row.email,
        "group": row.group_name,
        "weight": float(row.weight),
        "included": row.included,
        "selected": row.selected,
        "valid": not bool(row.validation_errors),
        "duplicate": row.duplicate_status,
        "errors": row.validation_errors,
    }


def roster_import_receipt_payload(receipt: RosterImportReceipt) -> dict:
    return {
        "id": str(receipt.pk),
        "mode": receipt.mode,
        "importedCount": receipt.imported_count,
        "createdCount": receipt.created_count,
        "updatedCount": receipt.updated_count,
        "resultsRevision": receipt.results_revision,
        "committedAt": receipt.committed_at.isoformat(),
    }
