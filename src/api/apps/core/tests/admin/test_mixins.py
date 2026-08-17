"""Tests for core admin mixins against the current application models."""

import io
import json
from datetime import date
from unittest.mock import patch

from django.contrib.admin.sites import AdminSite
from django.forms import Media
from django.http import QueryDict
from django.test import RequestFactory, TestCase
from django.utils import timezone
from openpyxl import load_workbook

from apps.core.admin.mixins import (
    DataExportMixin,
    ExcelExportMixin,
    TimestampedAdminMixin,
)
from apps.core.tests.helpers import make_admin, make_member
from apps.mail.models import EmailMessageLog
from apps.scheduling.models import Event

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _make_log(**kwargs):
    defaults = {
        "message_type": EmailMessageLog.MessageType.TEST,
        "recipient": "recipient@example.com",
        "subject": "Test message",
        "status": EmailMessageLog.Status.FAILED,
    }
    defaults.update(kwargs)
    return EmailMessageLog.objects.create(**defaults)


def _make_event(code="EXPORT1"):
    return Event.objects.create(
        organizer=make_member(email=f"{code.lower()}@example.com"),
        code=code,
        name="Export event",
        days=[1, 3, 5],
    )


class _MockSuperAdmin:
    """Minimal stand-in for ModelAdmin methods called by the mixins."""

    model = EmailMessageLog
    admin_site = AdminSite()

    @property
    def media(self):
        return Media()

    def get_queryset(self, request):
        return EmailMessageLog.objects.all()

    def get_list_display(self, request):
        return ["__str__"]

    def get_list_filter(self, request):
        return []

    def get_readonly_fields(self, request, obj=None):
        return []

    def get_actions(self, request, action_location=None):
        return {}

    def get_action(self, action, action_location=None):
        func = getattr(type(self), action)
        return func, action, getattr(func, "short_description", action)

    def save_model(self, request, obj, form, change):
        obj.save()


class _TimestampAdmin(TimestampedAdminMixin, _MockSuperAdmin):
    pass


class _DataExportAdmin(DataExportMixin, _MockSuperAdmin):
    pass


class _DataExportCustomAdmin(DataExportMixin, _MockSuperAdmin):
    export_fields = ["subject", "message_type", "created_at"]
    export_filename = "custom_email_logs"


def _load_xlsx(response_content):
    return load_workbook(io.BytesIO(response_content))


def _make_post_request(factory, data=None):
    request = factory.post("/admin/", data=data or {})
    request.META["SERVER_NAME"] = "testserver"
    return request


def _confirmed_post(factory, extra=None):
    data = {"action": "export_data", "export_confirm": "1", "export_format": "xlsx"}
    if extra:
        data.update(extra)
    return _make_post_request(factory, data)


class TimestampedAdminMixinTest(TestCase):
    def setUp(self):
        self.admin = _TimestampAdmin()
        self.request = RequestFactory().get("/admin/")

    def test_readonly_includes_timestamp_fields(self):
        readonly = self.admin.get_readonly_fields(self.request)
        self.assertIn("created_at", readonly)
        self.assertIn("updated_at", readonly)

    def test_list_display_includes_created_at(self):
        self.assertIn("created_at", self.admin.get_list_display(self.request))


class DataExportMixinTest(TestCase):
    def setUp(self):
        self.admin = _DataExportAdmin()
        self.factory = RequestFactory()
        self.request = self.factory.get("/admin/")

    def test_actions_include_export_data_as_unbound_method(self):
        actions = self.admin.get_actions(self.request)
        self.assertIn("export_data", actions)
        action = actions["export_data"]
        func = action.func if hasattr(action, "func") else action[0]
        self.assertFalse(hasattr(func, "__self__"))

    def test_actions_forward_location_and_omit_unavailable_export(self):
        action_location = object()
        actions = self.admin.get_actions(
            self.request,
            action_location=action_location,
        )
        self.assertIn("export_data", actions)

        with patch.object(self.admin, "get_action", return_value=None):
            actions = self.admin.get_actions(
                self.request,
                action_location=action_location,
            )
        self.assertNotIn("export_data", actions)

    def test_excel_export_mixin_alias(self):
        self.assertIs(ExcelExportMixin, DataExportMixin)

    def test_default_fields_use_current_model_fields_and_labels(self):
        fields = dict(self.admin.get_export_fields())
        self.assertEqual(list(fields), [field.name for field in EmailMessageLog._meta.fields])
        self.assertEqual(fields["message_type"], "Message Type")
        self.assertEqual(fields["provider_message_id"], "Provider Message Id")

    def test_backward_compat_get_excel_export_fields(self):
        self.assertEqual(self.admin.get_excel_export_fields(), self.admin.get_export_fields())

    def test_value_none_returns_empty_string(self):
        self.assertEqual(self.admin.get_export_value(_make_log(), "event"), "")

    def test_value_datetime_formatted(self):
        log = _make_log()
        dt = timezone.now()
        EmailMessageLog.objects.filter(pk=log.pk).update(created_at=dt)
        log.refresh_from_db()
        self.assertEqual(
            self.admin.get_export_value(log, "created_at"),
            dt.strftime("%Y-%m-%d %H:%M:%S"),
        )

    def test_value_uuid_returned_as_string(self):
        event = _make_event()
        admin = _DataExportAdmin()
        admin.model = Event
        self.assertEqual(admin.get_export_value(event, "event_id"), str(event.event_id))

    def test_value_string_returned_as_is(self):
        log = _make_log(subject="Hello")
        self.assertEqual(self.admin.get_export_value(log, "subject"), "Hello")

    def test_value_date_formatted(self):
        class FakeDateObject:
            date_field = date(2025, 3, 15)

        self.assertEqual(
            self.admin.get_export_value(FakeDateObject(), "date_field"),
            "2025-03-15",
        )

    def test_value_list_and_dict_serialized_as_json(self):
        log = _make_log()
        log.test_list = [{"q": "Color?", "a": "Blue"}]
        log.test_dict = {"key": "value", "num": 42}
        self.assertEqual(
            self.admin.get_export_value(log, "test_list"),
            '[{"q": "Color?", "a": "Blue"}]',
        )
        self.assertEqual(
            json.loads(self.admin.get_export_value(log, "test_dict")),
            {"key": "value", "num": 42},
        )

    def test_value_bool_returns_yes_no(self):
        class FakeBoolObject:
            active = True
            deleted = False

        self.assertEqual(self.admin.get_export_value(FakeBoolObject(), "active"), "Yes")
        self.assertEqual(self.admin.get_export_value(FakeBoolObject(), "deleted"), "No")

    def test_related_object_value_returned_as_str(self):
        event = _make_event("EXPORT2")
        log = _make_log(event=event)
        self.assertEqual(self.admin.get_export_value(log, "event"), str(event))

    def test_backward_compat_get_excel_export_value(self):
        log = _make_log(subject="Compatibility")
        self.assertEqual(
            self.admin.get_excel_export_value(log, "subject"),
            self.admin.get_export_value(log, "subject"),
        )

    def test_initial_post_renders_column_selection_template(self):
        log = _make_log()
        request = _make_post_request(
            self.factory,
            {"action": "export_data", "_selected_action": [str(log.pk)]},
        )
        request.user = make_admin(email="export-admin@example.com")

        response = self.admin.export_data(
            request,
            EmailMessageLog.objects.filter(pk=log.pk),
        )

        self.assertEqual(response.template_name, "admin/core/export_columns.html")
        self.assertIn(str(log.pk), response.context_data["pks"])
        self.assertGreater(len(response.context_data["available_fields"]), 0)
        self.assertEqual(response.context_data["default_filename"], "emailmessagelog")
        self.assertEqual(len(response.context_data["formats"]), 2)

    def test_xlsx_confirmed_post_returns_styled_workbook(self):
        first = _make_log(subject="Alpha", recipient="alpha@example.com")
        second = _make_log(subject="Beta", recipient="beta@example.com")
        queryset = EmailMessageLog.objects.filter(pk__in=[first.pk, second.pk]).order_by("subject")

        response = self.admin.export_data(_confirmed_post(self.factory), queryset)

        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("emailmessagelog_", response["Content-Disposition"])
        workbook = _load_xlsx(response.content)
        rows = list(workbook.active.iter_rows(values_only=True))
        header = rows[0]
        self.assertIn("Subject", header)
        self.assertIn("Recipient", header)
        self.assertEqual({row[header.index("Subject")] for row in rows[1:]}, {"Alpha", "Beta"})
        self.assertEqual(workbook.active.cell(row=1, column=1).fill.start_color.rgb, "004472C4")

    def test_selected_columns_limit_exported_fields(self):
        log = _make_log(subject="Selected", message_type="reminder")
        request = _make_post_request(self.factory, QueryDict(mutable=True))
        request.POST = QueryDict(mutable=True)
        request.POST.update(
            {"action": "export_data", "export_confirm": "1", "export_format": "xlsx"}
        )
        request.POST.setlist("export_fields", ["subject", "message_type"])

        response = self.admin.export_data(
            request,
            EmailMessageLog.objects.filter(pk=log.pk),
        )
        rows = list(_load_xlsx(response.content).active.iter_rows(values_only=True))
        self.assertEqual(rows, [("Subject", "Message Type"), ("Selected", "reminder")])

    def test_empty_selection_exports_all_columns(self):
        _make_log()
        response = self.admin.export_data(
            _confirmed_post(self.factory),
            EmailMessageLog.objects.all(),
        )
        header = next(_load_xlsx(response.content).active.iter_rows(values_only=True))
        self.assertEqual(len(header), len(EmailMessageLog._meta.fields))

    def test_user_filename_and_blank_fallback(self):
        _make_log()
        custom = self.admin.export_data(
            _confirmed_post(self.factory, {"export_filename": "my_custom_report"}),
            EmailMessageLog.objects.all(),
        )
        fallback = self.admin.export_data(
            _confirmed_post(self.factory, {"export_filename": "   "}),
            EmailMessageLog.objects.all(),
        )
        self.assertIn("my_custom_report_", custom["Content-Disposition"])
        self.assertIn("emailmessagelog_", fallback["Content-Disposition"])

    def test_json_format_returns_data_and_custom_filename(self):
        first = _make_log(subject="First", recipient="first@example.com")
        second = _make_log(subject="Second", recipient="second@example.com")
        response = self.admin.export_data(
            _confirmed_post(
                self.factory,
                {"export_format": "json", "export_filename": "message_report"},
            ),
            EmailMessageLog.objects.filter(pk__in=[first.pk, second.pk]),
        )

        self.assertEqual(response["Content-Type"], "application/json")
        self.assertIn("message_report_", response["Content-Disposition"])
        data = json.loads(response.content)
        self.assertEqual({item["subject"] for item in data}, {"First", "Second"})
        self.assertEqual(set(data[0]), {field.name for field in EmailMessageLog._meta.fields})

    def test_json_format_respects_column_selection(self):
        _make_log(subject="Filtered")
        request = _make_post_request(self.factory, QueryDict(mutable=True))
        request.POST = QueryDict(mutable=True)
        request.POST.update(
            {"action": "export_data", "export_confirm": "1", "export_format": "json"}
        )
        request.POST.setlist("export_fields", ["subject", "status"])
        response = self.admin.export_data(request, EmailMessageLog.objects.all())
        self.assertEqual(set(json.loads(response.content)[0]), {"subject", "status"})


class DataExportMixinCustomFieldsTest(TestCase):
    def setUp(self):
        self.admin = _DataExportCustomAdmin()
        self.factory = RequestFactory()

    def test_custom_fields_and_labels(self):
        self.assertEqual(
            self.admin.get_export_fields(),
            [
                ("subject", "Subject"),
                ("message_type", "Message Type"),
                ("created_at", "Created At"),
            ],
        )

    def test_custom_fields_and_filename_apply_to_xlsx(self):
        _make_log(subject="Custom", message_type="welcome")
        response = self.admin.export_data(
            _confirmed_post(self.factory),
            EmailMessageLog.objects.all(),
        )
        rows = list(_load_xlsx(response.content).active.iter_rows(values_only=True))
        self.assertIn("custom_email_logs_", response["Content-Disposition"])
        self.assertEqual(rows[0], ("Subject", "Message Type", "Created At"))
        self.assertEqual(rows[1][0:2], ("Custom", "welcome"))

    def test_custom_filename_applies_to_json(self):
        _make_log()
        response = self.admin.export_data(
            _confirmed_post(self.factory, {"export_format": "json"}),
            EmailMessageLog.objects.all(),
        )
        self.assertIn("custom_email_logs_", response["Content-Disposition"])

    def test_backward_compat_properties_and_setters(self):
        self.assertEqual(
            self.admin.excel_export_fields,
            ["subject", "message_type", "created_at"],
        )
        self.assertEqual(self.admin.excel_export_filename, "custom_email_logs")
        self.admin.excel_export_fields = ["subject"]
        self.admin.excel_export_filename = "renamed"
        self.assertEqual(self.admin.export_fields, ["subject"])
        self.assertEqual(self.admin.export_filename, "renamed")

    def test_get_export_fields_falls_back_for_unknown_field(self):
        self.admin.export_fields = ["subject", "not_a_real_field"]
        fields = dict(self.admin.get_export_fields())
        self.assertEqual(fields["subject"], "Subject")
        self.assertEqual(fields["not_a_real_field"], "Not A Real Field")


class DataExportCurrentSchedulingValueTest(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_json_export_with_date_value_does_not_crash(self):
        class FakeDateObject:
            date_field = date(2025, 6, 15)

        admin = _DataExportAdmin()
        admin.export_fields = ["date_field"]
        response = admin.export_data(
            _confirmed_post(self.factory, {"export_format": "json"}),
            [FakeDateObject()],
        )
        self.assertEqual(json.loads(response.content), [{"date_field": "2025-06-15"}])

    def test_json_export_with_current_json_field_does_not_crash(self):
        event = _make_event("EXPORT3")
        admin = _DataExportAdmin()
        admin.model = Event
        admin.export_fields = ["code", "days"]
        response = admin.export_data(
            _confirmed_post(self.factory, {"export_format": "json"}),
            Event.objects.filter(pk=event.pk),
        )
        self.assertEqual(json.loads(response.content), [{"code": "EXPORT3", "days": "[1, 3, 5]"}])
